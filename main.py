import flet as ft
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

import matplotlib as mpl
import time
import geometry_calculation
import volumes
import manufacturing_cost_plasma_cutting
import manufacturing_cost_drilling
import manufacturing_cost_lathe
import manufacturing_cost_root_pass_welding
import manufacturing_cost_welding
import manufacturing_cost_blade_saw_cutting
import raw_material_shell
import raw_material_flanges
import raw_material_tubes
import raw_material_tubesheets
import raw_material_nozzles
import raw_material_pass_partitions
import raw_material_tie_rods
import raw_material_removable_covers
import raw_material_nozzles_flanges
import raw_material_baffles
import raw_material_heads
import assembly_cost
import economic_model

#correr 150 ejemplos con la nueva version (test de consistencia) y a la nueva version se testea varias veces la novedad.
#al arreglar un bug. hay que identificar los ejemplos en los que el error ocurre. corriendo el nuevo 
#una vez superado el excel, tenerlo como repositorio de ejemplos.
#29/10/2024

#En el nuevo software, agregar unit testings para cada formula.
#En un excel, poner en cada columna el costos parciales, y su desglose

#Terminar el paper con un analisis de sensibilidad de cada parametro
#analizar la cantidad de parametros necesarios para obtener una desviacion del 5%

"""Tables used for calculating, should be moved later in next iterations
ask how to integrate them on the code, for now, we put it here 20/10/2024
"""
#Table D-5

table_d5 = [        # db        B        Rr        E 
{"db": 12.7,    "B": 31.75,     "Rr": 15.875,   "E":  15.875},#ok
{"db": 15.875,  "B": 38.1,      "Rr": 19.05,    "E":  19.05},   #ok
{"db": 19.05,   "B": 44.45,     "Rr": 20.6375,  "E":  20.6375}, #ok
{"db": 22.225,  "B": 52.3875,   "Rr": 23.8125,  "E":  23.8125}, #ok
{"db": 25.4,    "B": 57.15 ,    "Rr": 26.9875,  "E":  26.9875}, #ok,
{"db": 28.575,  "B": 63.5,      "Rr": 28.575,   "E":  28.575}, #o,k
{"db": 31.75,   "B": 71.4375,   "Rr": 31.75,    "E":  31.75},  #ok
{"db": 34.925,  "B": 77.7875,   "Rr": 34.925,   "E":  34.925}, #ok
{"db": 38.1,    "B": 82.55,     "Rr": 38.1,     "E":  38.1},  #}ok
{"db": 41.275,  "B": 88.9,      "Rr": 41.275,   "E":  41.275}, #ok
{"db": 44.45,   "B": 95.25,     "Rr": 44.45,    "E":  44.450}, #ok
{"db": 47.625,  "B": 101.6,     "Rr": 47.625,   "E":  47.625}, #ok
{"db": 50.8,    "B": 107.95,    "Rr": 50.8,     "E":  50.8}, #}ok
{"db": 57.15,   "B": 120.65,    "Rr": 57.15,    "E":  57.150}, #ok
{"db": 63.5,    "B": 133.35,    "Rr": 63.5,     "E":  60.325},#ok
{"db": 69.85,   "B": 146.05,    "Rr": 69.85,    "E":  66.675}, #ok
{"db": 76.2,    "B": 158.75,    "Rr": 76.2,     "E":  73.025}, #ok
{"db": 82.55,   "B": 168.275,   "Rr": 82.55,    "E":  76.2},  #}ok
{"db": 88.9,    "B": 180.975,   "Rr": 88.9,     "E":  82.55}, #ok
{"db": 95.25,   "B": 193.675,   "Rr": 95.25,    "E":  88.9}, #}ok
{"db": 101.6,   "B": 206.375,   "Rr": 101.6,    "E":  92.075}  #ok
 ]

#Table CB 4.41

table_cb441 = [
    {"Outside diameter shell": 356, "Distance between center baffles" : 305, "Baffle thickness": 1.6},
    {"Outside diameter shell": 356, "Distance between center baffles" : 610, "Baffle thickness": 3.2},
    {"Outside diameter shell": 356, "Distance between center baffles" : 914, "Baffle thickness": 4.8},
    {"Outside diameter shell": 356, "Distance between center baffles" : 1219, "Baffle thickness": 6.4},
    {"Outside diameter shell": 356, "Distance between center baffles" : 1524, "Baffle thickness": 9.5},
    {"Outside diameter shell": 356, "Distance between center baffles" : 100000, "Baffle thickness": 9.5},
    {"Outside diameter shell": 711, "Distance between center baffles" : 305, "Baffle thickness": 3.2},
    {"Outside diameter shell": 711, "Distance between center baffles" : 610, "Baffle thickness": 4.8},
    {"Outside diameter shell": 711, "Distance between center baffles" : 914, "Baffle thickness": 6.4},
    {"Outside diameter shell": 711, "Distance between center baffles" : 1219, "Baffle thickness": 9.5},
    {"Outside diameter shell": 711, "Distance between center baffles" : 1524, "Baffle thickness": 9.5},
    {"Outside diameter shell": 711, "Distance between center baffles" : 100000, "Baffle thickness": 12.7},
    {"Outside diameter shell": 965, "Distance between center baffles" : 305, "Baffle thickness": 4.8},
    {"Outside diameter shell": 965, "Distance between center baffles" : 610, "Baffle thickness": 6.4},
    {"Outside diameter shell": 965, "Distance between center baffles" : 914, "Baffle thickness": 7.9},
    {"Outside diameter shell": 965, "Distance between center baffles" : 1219, "Baffle thickness": 9.5},
    {"Outside diameter shell": 965, "Distance between center baffles" : 1524, "Baffle thickness": 12.7},
    {"Outside diameter shell": 965, "Distance between center baffles" : 100000, "Baffle thickness": 15.9},
    {"Outside diameter shell": 1524, "Distance between center baffles" : 305, "Baffle thickness": 6.4},
    {"Outside diameter shell": 1524, "Distance between center baffles" : 610, "Baffle thickness": 6.4},
    {"Outside diameter shell": 1524, "Distance between center baffles" : 914, "Baffle thickness": 9.5},
    {"Outside diameter shell": 1524, "Distance between center baffles" : 1219, "Baffle thickness": 12.7},
    {"Outside diameter shell": 1524, "Distance between center baffles" : 1524, "Baffle thickness": 15.9},
    {"Outside diameter shell": 1524, "Distance between center baffles" : 100000, "Baffle thickness": 15.9},
    {"Outside diameter shell": 2540, "Distance between center baffles" : 305, "Baffle thickness": 6.4},
    {"Outside diameter shell": 2540, "Distance between center baffles" : 610, "Baffle thickness": 9.5},
    {"Outside diameter shell": 2540, "Distance between center baffles" : 914, "Baffle thickness": 12.7},
    {"Outside diameter shell": 2540, "Distance between center baffles" : 1219, "Baffle thickness": 15.9},
    {"Outside diameter shell": 2540, "Distance between center baffles" : 1524, "Baffle thickness": 19.1},
    {"Outside diameter shell": 2540, "Distance between center baffles" : 100000, "Baffle thickness": 19.1}
    ]

#Table RCB 9.1.3.1
table_rcb9131 = [
        {"Diameter shell":610 , "Pass partition thickness":9.5},
        {"Diameter shell":1524, "Pass partition thickness":12.7},
        {"Diameter shell":2540, "Pass partition thickness":15.9}
        ]


#Table R 4.7.1

table_r471 = [
        {"Diameter shell": 381, "Tie rod diameter": 9.5, "Number of tie rods": 4},
        {"Diameter shell": 686, "Tie rod diameter": 9.5, "Number of tie rods": 6},
        {"Diameter shell": 838, "Tie rod diameter": 12.7, "Number of tie rods": 6},
        {"Diameter shell": 1219, "Tie rod diameter": 12.7, "Number of tie rods": 8},
        {"Diameter shell": 1524, "Tie rod diameter": 12.7, "Number of tie rods": 10},
        {"Diameter shell": 2540, "Tie rod diameter": 15.9, "Number of tie rods": 12}
        ]


#also stocks from should be moved to another place, for now, it'll stay here 20/10/2024
#A36 Steel Plates from https://www.metalsdepot.com/steel-products/steel-plate
    #Wp  Lp   tp           $
"""
plates_table = [
{"Width": 1524,"Length": 3048,"Thickness":    4.7625,"Price":     482.50},
{"Width": 1524,"Length": 3048,"Thickness":    6.35,"Price":       643.0  },
{"Width": 1524,"Length": 3048,"Thickness":    9.525,"Price":      1034.0},
{"Width": 1524,"Length": 3048,"Thickness":    12.7,"Price":       1470.0},
{"Width": 1219.2,"Length": 2438.4,"Thickness":15.875,"Price":    1544.00},
{"Width": 1524,"Length": 3048,"Thickness":     19,"Price":      2894.50},
{"Width": 1219.2,"Length": 2438.4,"Thickness":22.225,"Price": 2676.16},
{"Width": 1524,"Length": 3048,"Thickness":25.4,"Price":       3675.50},
{"Width": 1219.2,"Length": 2438.4,"Thickness":28.575,"Price":  6616.64 },
{"Width": 1219.2,"Length": 2438.4,"Thickness":31.75,"Price":   3822.40},
{"Width": 1219.2,"Length": 2438.4,"Thickness":34.925,"Price": 6468.48},
{"Width": 1219.2,"Length": 2438.4,"Thickness":37.7,"Price": 4586.88},
{"Width": 1219.2,"Length": 2438.4,"Thickness":44.45,"Price":5351.36},
{"Width": 1219.2,"Length": 2438.4,"Thickness":50.8,"Price":6116.16},
{"Width": 1219.2,"Length": 1219.2,"Thickness":63.5,"Price":3822.56}
]
"""
plates_table = [
{"Width": 3000,"Length": 10000,"Thickness":    4.7625,"Price":2243.14},
{"Width": 3000,"Length": 10000,"Thickness":    6.35,"Price":    2990.85},
{"Width": 3000,"Length": 10000,"Thickness":    9.525,"Price": 4468.28},
{"Width": 3000,"Length": 10000,"Thickness":    12.7,"Price":  5981.7},
{"Width": 3000,"Length": 10000,"Thickness":15.875,"Price":    7477.15},
{"Width": 3000,"Length": 10000,"Thickness":     19.05,"Price":   8972.55},
{"Width": 3000,"Length": 10000,"Thickness":22.225,"Price": 10467.98},
{"Width": 3000,"Length": 10000,"Thickness":25.4,"Price":   11963.4},
{"Width": 3000,"Length": 10000,"Thickness":38.1,"Price": 17945.1 },
{"Width": 3000,"Length": 10000,"Thickness":44.45,"Price":20935.95},
{"Width": 3000,"Length": 10000,"Thickness":50.8,"Price":23926.8},
{"Width": 3000,"Length": 10000,"Thickness":63.5,"Price":29908.5},
{"Width": 3000,"Length": 10000,"Thickness":69.85,"Price":32899.35},
{"Width": 3000,"Length": 10000,"Thickness":76.2,"Price":35890.2},
{"Width": 3000,"Length": 10000,"Thickness":101.6,"Price":47853.6}
]

#t
#Rods
    #OD    Lr   $
rods_table = [
{"Outside diameter": 3.175,"Length": 10000,"Price": 1.24},
{"Outside diameter": 6.35,"Length": 10000,"Price": 5.05},
{"Outside diameter": 9.525,"Length":10000,"Price": 11.20},
{"Outside diameter": 12.7,"Length": 10000,"Price":19.89},
{"Outside diameter": 15.875,"Length": 10000,"Price": 31.09},
{"Outside diameter": 19.05,"Length": 10000,"Price": 44.75},
{"Outside diameter": 22.225,"Length": 10000,"Price": 60.94},
{"Outside diameter": 25.4,"Length": 10000,"Price": 79.55},
{"Outside diameter": 28.575,"Length": 10000,"Price": 100.72},
{"Outside diameter": 31.75,"Length": 10000,"Price":124.3},
{"Outside diameter": 34.925,"Length": 10000,"Price": 150.45},
{"Outside diameter": 38.1,"Length": 10000,"Price":178.99},
{"Outside diameter": 41.275,"Length": 10000,"Price": 210.32}
]

#Tubes from SCHEDULE 40s and 40 https://www.trupply.com/products/seamless-pipe-a106b?variant=41891233464482
   #dte     L   t       $
schedule40_table = [
{"Outside diameter": 10.3,"Length":10000,"Wall thickness": 1.73,"Price": 7.31},
{"Outside diameter": 13.7,"Length":10000,"Wall thickness": 2.24,"Price": 12.66},
{"Outside diameter": 17.1,"Length":10000,"Wall thickness": 2.31,"Price": 16.85},
{"Outside diameter": 21.30,"Length":10000,"Wall thickness": 2.77,"Price": 25.32},
{"Outside diameter": 26.7,"Length":10000,"Wall thickness": 2.87,"Price":33.73},
{"Outside diameter": 33.4,"Length":10000,"Wall thickness": 3.38,"Price": 50.05},
{"Outside diameter": 42.2,"Length":10000,"Wall thickness": 3.56,"Price": 67.85},
{"Outside diameter": 48.3,"Length":10000,"Wall thickness": 3.68,"Price": 80.99},
{"Outside diameter": 60.3,"Length":10000,"Wall thickness": 3.91,"Price": 108.75},
{"Outside diameter": 73,"Length":10000,"Wall thickness": 5.16,"Price": 172.66},
{"Outside diameter": 88.9,"Length":10000,"Wall thickness": 5.49,"Price": 225.86},
{"Outside diameter": 101.6,"Length":10000,"Wall thickness": 5.74,"Price": 271.39},
{"Outside diameter": 114.3,"Length":10000,"Wall thickness": 6.02,"Price": 321.51},
{"Outside diameter": 141.3,"Length":10000,"Wall thickness": 6.55,"Price": 435.33},
{"Outside diameter": 168.3,"Length":10000,"Wall thickness": 7.11,"Price": 565.27},
{"Outside diameter": 219.1,"Length":10000,"Wall thickness": 8.18,"Price": 850.98},
{"Outside diameter": 273.0,"Length":10000,"Wall thickness": 9.27,"Price": 1205.84},
{"Outside diameter": 323.8,"Length":10000,"Wall thickness": 10.31,"Price": 1594.16},
{"Outside diameter": 355.6,"Length":10000,"Wall thickness": 11.13,"Price": 1891.02},
{"Outside diameter": 406.4,"Length":10000,"Wall thickness": 12.7,"Price": 2466.15},
{"Outside diameter": 457.0,"Length":10000,"Wall thickness": 14.27,"Price": 3116.11},
{"Outside diameter": 508.0,"Length":10000,"Wall thickness": 15.09,"Price": 3668.65},
{"Outside diameter": 610.0,"Length":10000,"Wall thickness": 17.48,"Price": 5108.51}
]

flanges_nozzles_table = [
        {"Inside diameter": 12.7, "Price": 11.16},
        {"Inside diameter": 19.05, "Price":12.2},
        {"Inside diameter": 25.4, "Price": 15.5},
        {"Inside diameter": 31.75, "Price": 24.29},
        {"Inside diameter": 38.1, "Price": 30.11},
        {"Inside diameter": 50.8, "Price": 30.02},
        {"Inside diameter": 63.5, "Price": 35.14},
        {"Inside diameter": 76.2, "Price": 45.15},
        {"Inside diameter": 106.6, "Price": 60.94},
        {"Inside diameter": 127, "Price": 86.4},
        {"Inside diameter": 152.4, "Price": 97.44},
        {"Inside diameter": 203.2, "Price": 153.8},
        {"Inside diameter": 254, "Price": 266.24},
        {"Inside diameter": 304.8, "Price": 428.48}
        ]

def heatmap(data, row_labels, col_labels, ax=None,
            cbar_kw=None, cbarlabel="", **kwargs):
    """
    Create a heatmap from a numpy array and two lists of labels.

    Parameters
    ----------
    data
        A 2D numpy array of shape (M, N).
    row_labels
        A list or array of length M with the labels for the rows.
    col_labels
        A list or array of length N with the labels for the columns.
    ax
        A `matplotlib.axes.Axes` instance to which the heatmap is plotted.  If
        not provided, use current Axes or create a new one.  Optional.
    cbar_kw
        A dictionary with arguments to `matplotlib.Figure.colorbar`.  Optional.
    cbarlabel
        The label for the colorbar.  Optional.
    **kwargs
        All other arguments are forwarded to `imshow`.
    """

    if ax is None:
        ax = plt.gca()

    if cbar_kw is None:
        cbar_kw = {}

    # Plot the heatmap
    im = ax.imshow(data, **kwargs)

    # Create colorbar
    cbar = ax.figure.colorbar(im, ax=ax, **cbar_kw)
    cbar.ax.set_ylabel(cbarlabel, rotation=90, va="top")

    # Show all ticks and label them with the respective list entries.
    ax.set_xticks(range(data.shape[1]), labels=col_labels,
                  rotation=-30, ha="right", rotation_mode="anchor")
    ax.set_yticks(range(data.shape[0]), labels=row_labels)

    # Let the horizontal axes labeling appear on top.
    ax.tick_params(top=True, bottom=False,
                   labeltop=True, labelbottom=False)

    # Turn spines off and create white grid.
    ax.spines[:].set_visible(False)

    ax.set_xticks(np.arange(data.shape[1]+1)-.5, minor=True)
    ax.set_yticks(np.arange(data.shape[0]+1)-.5, minor=True)
    ax.grid(which="minor", color="w", linestyle='-', linewidth=3)
    ax.tick_params(which="minor", bottom=False, left=False)

    return im, cbar

def main(page: ft.Page):

    page.title = "Heat Exchanger manufacturing cost"
    page.vertical_alignment = ft.alignment.center
    page.horizontal_alignment = ft.alignment.center
    page.scroll = "adaptive"
    global raw_material


    def total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting,plates_table,schedule40_table,rods_table): 
        global ts,diameter_bolts,diameter_flanges, flanges_thickness, number_bolts, tubesheets_thickness, tubes_length, baffles_diameter, baffles_thickness, shell_nozzles_inside_diameter, tubes_nozzles_inside_diameter, shell_side_nozzles_length, tubes_side_nozzles_length, shell_side_nozzles_thickness, tubes_side_nozzles_thickness, heads_length, heads_diameter, heads_thickness, pass_partitions_length, pass_partitions_width, pass_partitions_length, pass_partitions_thickness, removable_covers_diameter, removable_covers_thickness, diameter_tie_rods, length_tie_rods, number_tie_rods, tubes_thickness, tubesheets_diameter


        ts = geometry_calculation.calculate_shell_thickness(shell_side_pressure, tube_side_pressure, ds.value, max_allowable_stress.value, welding_efficiency.value)

        diameter_bolts= geometry_calculation.calculate_flange_and_bolts(table_d5, ds, ts,shell_side_pressure, tube_side_pressure, gasket_material_factor, max_allowable_stress, Sb, welding_efficiency)[0] 

        diameter_flanges= geometry_calculation.calculate_flange_and_bolts(table_d5, ds, ts,shell_side_pressure, tube_side_pressure, gasket_material_factor, max_allowable_stress, Sb, welding_efficiency)[1] 

        flanges_thickness= geometry_calculation.calculate_flange_and_bolts(table_d5, ds, ts,shell_side_pressure, tube_side_pressure, gasket_material_factor, max_allowable_stress, Sb, welding_efficiency)[2]

        number_bolts = geometry_calculation.calculate_flange_and_bolts(table_d5, ds, ts,shell_side_pressure, tube_side_pressure, gasket_material_factor, max_allowable_stress, Sb, welding_efficiency)[3]
        print(f"Nb is {Nb}")
        
        tubesheets_thickness = geometry_calculation.calculate_tubesheets_thickness(plates_table, ds, shell_side_pressure, tube_side_pressure, corrosion_allowance, max_allowable_stress, lay, ltp, dte, diameter_flanges)[0]

        tubesheets_diameter = geometry_calculation.calculate_tubesheets_thickness(plates_table, ds, shell_side_pressure, tube_side_pressure, corrosion_allowance, max_allowable_stress, lay, ltp, dte, diameter_flanges)[1]


        tubes_length = geometry_calculation.calculate_tubes_length(ls, tubesheets_thickness)
        tubes_thickness = geometry_calculation.calculate_tubes_thickness(schedule40_table,dte)
        
        baffles_diameter = geometry_calculation.calculate_baffles_diameter(ds)

        baffles_thickness = geometry_calculation.calculate_baffles_thickness(table_cb441, ds, ls, Nb)
        
        shell_nozzles_inside_diameter = geometry_calculation.calculate_nozzles_inside_diameters(shell_side_fluid_density, tube_side_fluid_density, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, schedule40_table)[0]

        tubes_nozzles_inside_diameter = geometry_calculation.calculate_nozzles_inside_diameters(shell_side_fluid_density, tube_side_fluid_density, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, schedule40_table)[1]


        shell_side_nozzles_length = geometry_calculation.calculate_nozzles_length_thickness(diameter_flanges, ds, shell_nozzles_inside_diameter, tubes_nozzles_inside_diameter, schedule40_table)[0]
        
        tubes_side_nozzles_length = geometry_calculation.calculate_nozzles_length_thickness(diameter_flanges, ds, shell_nozzles_inside_diameter, tubes_nozzles_inside_diameter, schedule40_table)[2]

        shell_side_nozzles_thickness = geometry_calculation.calculate_nozzles_length_thickness(diameter_flanges, ds, shell_nozzles_inside_diameter, tubes_nozzles_inside_diameter, schedule40_table)[1]

        tubes_side_nozzles_thickness = geometry_calculation.calculate_nozzles_length_thickness(diameter_flanges, ds, shell_nozzles_inside_diameter, tubes_nozzles_inside_diameter, schedule40_table)[3]

        heads_length, heads_thickness, heads_diameter = geometry_calculation.calculate_heads(ds, ts, tubes_nozzles_inside_diameter)

        pass_partitions_width = geometry_calculation.calculate_pass_partitions(heads_diameter, heads_length, table_rcb9131, ds)[0]

        pass_partitions_length = geometry_calculation.calculate_pass_partitions(heads_diameter, heads_length, table_rcb9131, ds)[1]

        pass_partitions_thickness = geometry_calculation.calculate_pass_partitions(heads_diameter, heads_length, table_rcb9131, ds)[2]

        removable_covers_diameter, removable_covers_thickness = geometry_calculation.calculate_removable_covers(diameter_flanges, flanges_thickness)

        number_tie_rods, diameter_tie_rods, length_tie_rods = geometry_calculation.calculate_tie_rods(ds, table_r471, ls)



        global volume_shell,volume_tubes,volume_tubesheets,volume_baffles,volume_heads,volume_shell_side_nozzles,volume_tubes_side_nozzles,volume_removable_covers,volume_flange,volume_pass_partitions,volume_tie_rods , volume_heat_exchanger

        volume_shell = volumes.volume_shell(ds, ls, ts, shell_nozzles_inside_diameter)
        volume_tubes = volumes.volume_tubes(tubes_length, Ntp, number_passes, dte, tubes_thickness)

        volume_tubesheets = volumes.volume_tubesheets(tubesheets_thickness, tubesheets_diameter, Ntp, number_passes, dte, number_bolts, diameter_bolts)
        
        volume_baffles = volumes.volume_baffles(baffles_diameter, Ntp, number_passes, BC, dte, Nb, baffles_thickness, diameter_tie_rods, number_tie_rods)

        volume_heads = volumes.volume_heads(heads_length, ts, ds, tubes_nozzles_inside_diameter)

        volume_shell_side_nozzles = volumes.volume_shell_side_nozzles(shell_side_nozzles_length,shell_side_nozzles_thickness, shell_nozzles_inside_diameter)

        volume_tubes_side_nozzles = volumes.volume_tubes_side_nozzles(tubes_side_nozzles_length,tubes_side_nozzles_thickness, tubes_nozzles_inside_diameter)

        volume_removable_covers = volumes.volume_removable_cover(removable_covers_diameter, removable_covers_thickness, number_bolts, diameter_bolts)

        volume_flange = volumes.volume_flange(ds, ts, diameter_flanges, flanges_thickness)

        volume_pass_partitions = volumes.volume_pass_partitions(number_passes, ds, pass_partitions_thickness, heads_length)

        volume_tie_rods = volumes.volume_tie_rods(number_tie_rods, diameter_tie_rods, length_tie_rods)
        
        volume_heat_exchanger = volume_shell+ volume_tubes+volume_tubesheets + volume_baffles + volume_heads + volume_shell_side_nozzles + volume_tubes_side_nozzles + volume_removable_covers + volume_flange + volume_pass_partitions + volume_tie_rods #add volume of flange nozzles later
        TOTAL_WEIGHT = volume_heat_exchanger * 7850 
        print(f"Total weight of HEX: {TOTAL_WEIGHT} [kg]")

        global raw_material 
        global cost_raw_material_shell
        global cost_raw_material_tubes
        global cost_raw_material_tubesheets
        global cost_raw_material_heads
        global cost_raw_material_nozzles
        global cost_raw_material_flanges
        global cost_raw_material_removable_covers
        global cost_raw_material_pass_partitions
        global cost_raw_material_tie_rods
        global cost_raw_material_nozzles_flanges

        global number_plates_shell, number_pipes_shell, Ntrunks
        
        cost_raw_material_shell = cost_raw_material_shell = raw_material_shell.cost_raw_material_shell(ds, schedule40_table, ls, ts, plates_table)[0]

        number_plates_shell = raw_material_shell.cost_raw_material_shell(ds, schedule40_table, ls, ts, plates_table)[1]

        number_pipes_shell = raw_material_shell.cost_raw_material_shell(ds, schedule40_table, ls, ts, plates_table)[2]

        Ntrunks = raw_material_shell.cost_raw_material_shell(ds, schedule40_table, ls, ts, plates_table)[3]

        scrap_material_shell = raw_material_shell.scrap_material_shell(plates_table,
                                                                       schedule40_table,
                                                                       ds,
                                                                       ts,
                                                                       ls,
                                                                       number_plates_shell,
                                                                       number_pipes_shell)
        cost_raw_material_tubesheets = raw_material_tubesheets.cost_raw_material_tubesheets(tubesheets_thickness, tubesheets_diameter, plates_table, ec)[0]
        number_plates_tubesheets = raw_material_tubesheets.cost_raw_material_tubesheets(tubesheets_thickness, tubesheets_diameter, plates_table, ec)[1]
    
        scrap_material_tubesheets = raw_material_tubesheets.scrap_material_tubesheets(plates_table,tubesheets_thickness, tubesheets_diameter, number_plates_tubesheets)

        cost_raw_material_tubes = raw_material_tubes.cost_raw_material_tubes(tubes_length,
                                                                                dte,
                                                                                Ntp,
                                                                                number_passes,
                                                                                schedule40_table)[0]
        number_tubes_tubes = raw_material_tubes.cost_raw_material_tubes(tubes_length,
                                                                                dte,
                                                                                Ntp,
                                                                                number_passes,
                                                                                schedule40_table)[1]
    
        scrap_material_tubes = raw_material_tubes.scrap_material_tubes(dte,
                                                                        schedule40_table,
                                                                    number_tubes_tubes,
                                                                    Ntp,
                                                                        number_passes,
                                                                        tubes_length)
    
        cost_raw_material_heads = raw_material_heads.cost_raw_material_heads(heads_diameter,heads_thickness, heads_length,schedule40_table, plates_table)[0]
        number_pipes_heads = raw_material_heads.cost_raw_material_heads(heads_diameter,
                                                                        heads_thickness,
                                                        heads_length,
                                                        schedule40_table,
                                                        plates_table)[1]
        number_plates_heads = raw_material_heads.cost_raw_material_heads(heads_diameter,
                                                                            heads_thickness,
                                                        heads_length,
                                                        schedule40_table,
                                                        plates_table)[2]
    
        scrap_material_heads = raw_material_heads.scrap_material_heads(heads_diameter,
                                                                        heads_thickness,
                                                                        heads_length,
                                                                        schedule40_table,
                                                                        plates_table)
            
        cost_raw_material_nozzles = raw_material_nozzles.cost_raw_material_nozzles(shell_nozzles_inside_diameter, tubes_nozzles_inside_diameter, schedule40_table)[0]
        number_pipes_nozzles = raw_material_nozzles.cost_raw_material_nozzles(shell_nozzles_inside_diameter, tubes_nozzles_inside_diameter, schedule40_table)[1]
    
        scrap_material_shell_nozzles = raw_material_nozzles.scrap_material_shell_nozzles(shell_nozzles_inside_diameter,
                                                                                            shell_side_nozzles_thickness,
                                                                                            shell_side_nozzles_length,
                                                                                            schedule40_table)
    
    
        scrap_material_tubes_nozzles = raw_material_nozzles.scrap_material_tubes_nozzles(tubes_nozzles_inside_diameter,
                                                                                            tubes_side_nozzles_thickness,
                                                                                            tubes_side_nozzles_length,
                                                                                            schedule40_table)
    
    
    
        cost_raw_material_baffles = raw_material_baffles.cost_raw_material_baffles(baffles_thickness,
                                                                                Nb,
                                                                                ec,
                                                                                BC,
                                                                                baffles_diameter,
                                                                                plates_table)[0]
    
        number_plates_baffles = raw_material_baffles.cost_raw_material_baffles(baffles_thickness,
                                                                                Nb,
                                                                                    ec,
                                                                                    BC,
                                                                                    baffles_diameter,
                                                                                    plates_table)[1]
    
        scrap_material_baffles = raw_material_baffles.scrap_material_baffles(baffles_diameter,
                                                                                BC,
                                                                                Nb,
                                                                                baffles_thickness, 
                                                                                number_plates_baffles,
                                                                                plates_table)
    
    
            
        cost_raw_material_flanges = raw_material_flanges.cost_raw_material_flanges(diameter_flanges,
                                                                                    ec,
                                                                                    flanges_thickness,
                                                                                    plates_table)[0]
    
        number_plates_flanges = raw_material_flanges.cost_raw_material_flanges(diameter_flanges,
                                                                                    ec,
                                                                                    flanges_thickness,
                                                                                    plates_table)[1]
    
        scrap_material_flanges = raw_material_flanges.scrap_material_flanges(diameter_flanges, ds, ts, flanges_thickness, ec, plates_table, number_plates_flanges)
    
    
        cost_raw_material_removable_covers = raw_material_removable_covers.cost_raw_material_removable_covers(removable_covers_thickness, ec, removable_covers_diameter, plates_table)[0]
    
        number_plates_removable_covers = raw_material_removable_covers.cost_raw_material_removable_covers(removable_covers_thickness, ec, removable_covers_diameter, plates_table)[1]
    
        scrap_material_removable_covers = raw_material_removable_covers.scrap_material_removable_covers(removable_covers_thickness, plates_table, removable_covers_diameter, number_plates_removable_covers)  

    
        cost_raw_material_pass_partitions = raw_material_pass_partitions.cost_raw_material_pass_partitions(pass_partitions_thickness, plates_table)[0]
            
        number_plates_pass_partitions = raw_material_pass_partitions.cost_raw_material_pass_partitions(pass_partitions_thickness, plates_table)[1]
    
        scrap_material_pass_partitions = raw_material_pass_partitions.scrap_material_pass_partitions(pass_partitions_thickness, pass_partitions_width, pass_partitions_length, number_passes, plates_table)
    
        cost_raw_material_tie_rods = raw_material_tie_rods.cost_raw_material_tie_rods(number_tie_rods,
                                                                                        length_tie_rods,
                                                                                        diameter_tie_rods,
                                                                                        rods_table)[0]
    
        number_rods_tie_rods = raw_material_tie_rods.cost_raw_material_tie_rods(number_tie_rods,
                                                                                        length_tie_rods,
                                                                                        diameter_tie_rods,
                                                                                        rods_table)[1]
    
        scrap_material_tie_rods = raw_material_tie_rods.scrap_material_tie_rods(number_rods_tie_rods,
                                                                                length_tie_rods,
                                                                                number_tie_rods,
                                                                                diameter_tie_rods,
                                                                                rods_table)
    
        cost_raw_material_nozzles_flanges = raw_material_nozzles_flanges.cost_raw_material_nozzles_flanges(flanges_nozzles_table,
                                                                                                            shell_nozzles_inside_diameter,
                                                                                                            tubes_nozzles_inside_diameter)[0]
    
        number_nozzles_flanges = raw_material_nozzles_flanges.cost_raw_material_nozzles_flanges(flanges_nozzles_table,
                                                                                                            shell_nozzles_inside_diameter,
                                                                                                           tubes_nozzles_inside_diameter)[1]

        global raw_material
        raw_material = [
                {"Name of part": "Shell","K scrap factor": k_scrap_factor_shell,"Cost": cost_raw_material_shell, "Scrap": scrap_material_shell},
                {"Name of part": "Tubes","K scrap factor": k_scrap_factor_tubes,"Cost": cost_raw_material_tubes, "Scrap": scrap_material_tubes},
                {"Name of part": "Tubesheets","K scrap factor": k_scrap_factor_tubesheets, "Cost": cost_raw_material_tubesheets, "Scrap": scrap_material_tubesheets},
                {"Name of part": "Heads","K scrap factor": k_scrap_factor_heads,"Cost": cost_raw_material_heads, "Scrap": scrap_material_heads},
                {"Name of part": "Shell nozzles","K scrap factor": k_scrap_factor_shell_nozzles,"Cost": cost_raw_material_nozzles/2, "Scrap": scrap_material_shell_nozzles},
                {"Name of part": "Tubes nozzles","K scrap factor": k_scrap_factor_tubes_nozzles,"Cost": cost_raw_material_nozzles/2, "Scrap": scrap_material_tubes_nozzles},
                {"Name of part": "Flanges","K scrap factor": k_scrap_factor_flanges,"Cost": cost_raw_material_flanges, "Scrap": scrap_material_flanges},
                {"Name of part": "Removable covers","K scrap factor": k_scrap_factor_removable_covers,"Cost": cost_raw_material_removable_covers, "Scrap": scrap_material_removable_covers},
                {"Name of part": "Pass partitions","K scrap factor": k_scrap_factor_pass_partitions,"Cost": cost_raw_material_pass_partitions, "Scrap": scrap_material_pass_partitions},
                {"Name of part": "Tie rods","K scrap factor": k_scrap_factor_tie_rods,"Cost": cost_raw_material_tie_rods, "Scrap": scrap_material_tie_rods},
                {"Name of part": "Nozzles flanges","K scrap factor": k_scrap_factor_nozzles_flanges,"Cost": cost_raw_material_nozzles_flanges, "Scrap": 0}
                ]



        global cost_assembly
        cost_assembly = assembly_cost.cost_assembly(labor_cost_assembly,
                                                    starting_time_assembly,
                                                    manipulation_time_assembly,
                                                    volume_heat_exchanger,
                                                    labor_cost_bolted_joints,
                                                    material_cost_bolted_joints,
                                                    utility_cost_bolted_joints,
                                                    depreciation_cost_bolted_joints,
                                                    number_bolts,
                                                    STba)
        print(f"cost of assembly is {cost_assembly}")
        

#Now calculate manufacturing cost of everything
        global cost_plasma_cutting, cost_drilling, cost_lathe, cost_root_pass_welding, cost_welding, cost_blade_saw_cutting

        cost_plasma_cutting = manufacturing_cost_plasma_cutting.calculate_cost_plasma_cutting(labor_cost_plasma_cutting, 
                                                                                              material_cost_plasma_cutting, 
                                                                                              utility_cost_plasma_cutting, 
                                                                                              depreciation_cost_plasma_cutting, 
                                                                                              starting_time_plasma_cutting, 
                                                                                              manipulation_time_plasma_cutting, 
                                                                                              spca, 
                                                                                              tpcs, 
                                                                                              operation_factor_plasma_cutting, 
                                                                                              ec, 
                                                                                              ts, 
                                                                                              flanges_thickness, 
                                                                                              baffles_thickness, 
                                                                                              pass_partitions_thickness, 
                                                                                              removable_covers_thickness, 
                                                                                              heads_thickness, 
                                                                                              tubesheets_thickness, 
                                                                                              diameter_flanges, 
                                                                                              ds, 
                                                                                              pass_partitions_width, 
                                                                                              pass_partitions_length, 
                                                                                              number_passes, 
                                                                                              removable_covers_diameter, 
                                                                                              heads_diameter, 
                                                                                              tubesheets_diameter, 
                                                                                              heads_length, 
                                                                                              Nb,  
                                                                                              volume_tubesheets, 
                                                                                              volume_flange, 
                                                                                              volume_baffles, 
                                                                                              volume_pass_partitions, 
                                                                                              volume_removable_covers, 
                                                                                              volume_shell, 
                                                                                              volume_heads, 
                                                                                              ls)

        cost_drilling = manufacturing_cost_drilling.calculate_cost_drilling(labor_cost_drilling, 
                                                                            material_cost_drilling, 
                                                                            utility_cost_drilling, 
                                                                            depreciation_cost_drilling, 
                                                                            starting_time_drilling, 
                                                                            manipulation_time_drilling, 
                                                                            Stm, 
                                                                            rpm, 
                                                                            fnd, 
                                                                            STdo, 
                                                                            SDh, 
                                                                            flanges_thickness, 
                                                                            baffles_thickness,
                                                                            removable_covers_thickness, 
                                                                            tubesheets_thickness, 
                                                                            Ntp, 
                                                                            number_passes, 
                                                                            Nb, 
                                                                            number_tie_rods, 
                                                                            volume_tubesheets, 
                                                                            volume_flange, 
                                                                            volume_baffles, 
                                                                            volume_pass_partitions, 
                                                                            volume_removable_covers, 
                                                                            volume_shell, 
                                                                            volume_heads, 
                                                                            dte, 
                                                                            number_bolts, 
                                                                            diameter_bolts, 
                                                                            diameter_tie_rods)


        cost_lathe = manufacturing_cost_lathe.calculate_cost_lathe(labor_cost_lathe,
                                                                   material_cost_lathe,
                         utility_cost_lathe,
                         depreciation_cost_lathe,
                         starting_time_lathe,
                         manipulation_time_lathe,
                         operation_factor_lathe,
                         vcl,
                         apl, 
                         fnl, 
                         diameter_flanges, 
                         ds,
                                                                   tubesheets_diameter,
                         flanges_thickness, 
                         removable_covers_thickness, 
                         tubesheets_thickness, 
                         volume_tubesheets, 
                         volume_flange,
                         volume_removable_covers,
                                                                   removable_covers_diameter,
                                                                   ts)


        cost_root_pass_welding = manufacturing_cost_root_pass_welding.calculate_cost_root_pass_welding(labor_cost_root_pass_welding,
                                material_cost_root_pass_welding, 
                                utility_cost_root_pass_welding,
                                depreciation_cost_root_pass_welding, 
                                starting_time_root_pass_welding, 
                                manipulation_time_root_pass_welding, 
                                density_filler_material, 
                                rrp , 
                                operation_factor_root_pass_welding, 
                                nrp,  
                                                                                                  thickness_root_pass_welding,
                                ts, 
                                flanges_thickness, 
                                baffles_thickness, 
                                pass_partitions_thickness, 
                                removable_covers_thickness, 
                                heads_thickness, 
                                tubesheets_thickness, 
                                number_plates_shell,
                                                                                                       Ntrunks,
                                diameter_flanges, 
                                ds, 
                                pass_partitions_width, 
                                pass_partitions_length,
                                number_passes, 
                                Ntp, 
                                tubes_thickness, 
                                removable_covers_diameter, 
                                heads_diameter, 
                                tubesheets_diameter, 
                                heads_length, 
                                Nb,  
                                volume_tubesheets, 
                                volume_flange, 
                                volume_baffles, 
                                volume_pass_partitions, 
                                volume_removable_covers, 
                                volume_shell, 
                                volume_heads,
                                                                                                  volume_shell_side_nozzles,
                                                                                                       volume_tubes_side_nozzles,
                                                                                                  shell_nozzles_inside_diameter,
                                                                                                       tubes_nozzles_inside_diameter,
                                                                                                  dte,
                                                                                                  volume_tubes,
                                                                                                  number_tie_rods,
                                                                                                  diameter_tie_rods,
                                                                                                  volume_tie_rods,
                                                                                                  ls,
                                                                                                  tubes_length)
        

        cost_welding = manufacturing_cost_welding.calculate_cost_welding(labor_cost_welding, 
                                material_cost_welding, 
                                utility_cost_welding,
                                depreciation_cost_welding, 
                                starting_time_welding, 
                                manipulation_time_welding, 
                                density_filler_material, 
                                rw , 
                                operation_factor_welding, 
                                nw,  
                                                                                                  thickness_root_pass_welding,
                                ts, 
                                flanges_thickness, 
                                baffles_thickness, 
                                pass_partitions_thickness, 
                                removable_covers_thickness, 
                                heads_thickness, 
                                tubesheets_thickness, 
                                number_plates_shell,
                                                                                                       Ntrunks,
                                diameter_flanges, 
                                ds, 
                                pass_partitions_width, 
                                pass_partitions_length,
                                number_passes, 
                                Ntp, 
                                tubes_thickness, 
                                removable_covers_diameter, 
                                heads_diameter, 
                                tubesheets_diameter, 
                                heads_length, 
                                Nb,  
                                volume_tubesheets, 
                                volume_flange, 
                                volume_baffles, 
                                volume_pass_partitions, 
                                volume_removable_covers, 
                                volume_shell, 
                                volume_heads,
                                                                                                  volume_shell_side_nozzles,
                                                                                                       volume_tubes_side_nozzles,
                                                                                                  shell_nozzles_inside_diameter,
                                                                                                       tubes_nozzles_inside_diameter,
                                                                                                  dte,
                                                                                                  volume_tubes,
                                                                                                  number_tie_rods,
                                                                                                  diameter_tie_rods,
                                                                                                  volume_tie_rods,
                                                                                                  ls,
                                                                                                  tubes_length,shell_side_nozzles_thickness,tubes_side_nozzles_thickness)
        

        cost_blade_saw_cutting = manufacturing_cost_blade_saw_cutting.calculate_cost_blade_saw_cutting(labor_cost_blade_saw_cutting, 
                                     material_cost_blade_saw_cutting, 
                                     utility_cost_blade_saw_cutting, 
                                     depreciation_cost_blade_saw_cutting, 
                                     starting_time_blade_saw_cutting, 
                                     manipulation_time_blade_saw_cutting, 
                                     operation_factor_blade_saw_cutting, 
                                     shell_nozzles_inside_diameter,
                                     tubes_nozzles_inside_diameter,
                                     shell_side_nozzles_thickness,
                                     tubes_side_nozzles_thickness, 
                                     dte, 
                                     tubes_thickness, 
                                     Crbs, 
                                     Ntp, 
                                     number_passes,
                                     volume_tubes,
                                     volume_shell_side_nozzles,
                                     volume_tubes_side_nozzles)

        total_cost_heat_exchanger_FOB = economic_model.cost_heat_exchanger_FOB(raw_material,
                                                                               cost_plasma_cutting,
                                                                               cost_drilling,
                                                                               cost_lathe,
                                                                               cost_root_pass_welding,
                                                                               cost_welding,
                                                                               cost_blade_saw_cutting,
                                                                               cost_assembly,
                                                                               kg_price
                                                                               )
        return total_cost_heat_exchanger_FOB, raw_material, cost_raw_material_shell,cost_raw_material_tubesheets,cost_raw_material_tubes,cost_raw_material_heads, cost_raw_material_nozzles,cost_raw_material_flanges,cost_raw_material_removable_covers,cost_raw_material_pass_partitions,cost_raw_material_tie_rods,cost_raw_material_nozzles_flanges,cost_raw_material_baffles
        
    
    def calculate_heat_exchanger_FOB_button(e):
        print(f"\n\n\n ##### lets see is ds is available: {ds.value}")
        total_cost = total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]
        print(f"\n \n Cost of manufaturing Plasma Cutting: ${cost_plasma_cutting} \n")
        print(f"\n \n Cost of manufaturing Drilling: ${cost_drilling} \n")
        print(f"\n \n Cost of manufaturing Lathe: ${cost_lathe} \n")
        print(f"\n \n Cost of manufaturing Root pass Welding: ${cost_root_pass_welding} \n")
        print(f"\n \n Cost of manufaturing Welding: ${cost_welding} \n")
        print(f"\n \n Cost of manufaturing Blade Saw Cutting: ${cost_blade_saw_cutting} \n")

        data_entry.value =f"the total cost of heat exchanger FOB is {total_cost}$"
        page.update()




#########################
# SENSITIVITY ANALYSIS # 
########################
    def sensitivity_analysis(parameter, sensitivity_amount):
        original_value_parameter = parameter.value
        print(parameter)
        if parameter.value==0:
            h=1
        else:
            h = parameter.value * sensitivity_amount #amount of change in parameter, sensityvity_amount shoulde be 0.25 or 0.5
#from here on, calculates onlye sensitivity parameter
        total_cost_heat_exchanger_FOB_before = total_cost_heat_exchanger_FOB()
        factor_p_over_y = parameter.value/total_cost_heat_exchanger_FOB_before

        print(f"Factor p over y is {factor_p_over_y}")
        parameter.value += h  #increase the parameter in an amount h
        total_cost_heat_exchanger_FOB_after = total_cost_heat_exchanger_FOB()
        sensitivity = factor_p_over_y * ((total_cost_heat_exchanger_FOB_after-total_cost_heat_exchanger_FOB_before)/h)
#up to h, sensitivity is definded
        parameter.value= original_value_parameter #parameter goes back to original value
        return sensitivity


##################################
# HEAT MAP FOR ALL EXCHANGERS   #
################################







#Data Entry
    global number_passes,ds,ls,dte
    ds = ft.TextField(label="Shell inside diameter", width=200, value=1219) #agregar unidades de todo
    ls = ft.TextField(label="Shell length", width=200, value=4876.8)
    dte = ft.TextField(label="Outside diameter of tubes", width=200, value=12.7)
    lay = ft.TextField(label="Pitch type", width=200, value="Triangle")
    ltp = ft.TextField(label="Tube pitch ratio", width=200, value=1.25)
    Nb = ft.TextField(label="Number of baffles", width=200, value=14)
    BC = ft.TextField(label="Baffle cut", width=200, value=25)
    number_passes = ft.TextField(label="Number of passes", width=200, value=4)
    Ntp = ft.TextField(label="Number of tubes per pass", width=200, value=449)
    
    #Parameters
    max_allowable_stress = ft.TextField(label="Maximum Allowable Stress", value=120)
    Sb = ft.TextField(label="Maximum Allowable Bolt Stress", value=120)
    gasket_material_factor = ft.TextField(label="Gasket Material Factor", value=1.25)
    welding_efficiency = ft.TextField(label="Welding efficiency", value=0.8)
    corrosion_allowance = ft.TextField(label="Corrosion allowance", value=16)

    #Fluid Parameters
    shell_side_pressure = ft.TextField(label="Shell Side Pressure", value=0.2)
    tube_side_pressure = ft.TextField(label="Tube Side pressure", value=0.2)
    shell_side_fluid_velocity = ft.TextField(label="Shell Side Fluid Velocoity", value=2)
    tube_side_fluid_velocity = ft.TextField(label="Tube Side Fluid Velocity", value=2)
    shell_side_mass_flow = ft.TextField(label="Shell Side Mass Flow", value=110)
    tube_side_mass_flow = ft.TextField(label="Tube Side Mass Flow", value=130)
    shell_side_fluid_density = ft.TextField(label="Shell Side Fluid Density", value=786)
    tube_side_fluid_density = ft.TextField(label="Tube Side Fluid Density", value=995)
  

    #Assembly Settings
    labor_cost_assembly = ft.TextField(label="Labor cost per hour of assembly", value=19.968)
    starting_time_assembly= ft.TextField(label="starting time assembly", value=0.02)
    manipulation_time_assembly = ft.TextField(label="Manipulation Time per 1000kg of material", value=1)
    
    labor_cost_bolted_joints = ft.TextField(label="Labor cost per hour of bolted joints", value=13)
    material_cost_bolted_joints = ft.TextField(label="Material cost per hour of bolted joints", value=0)
    utility_cost_bolted_joints = ft.TextField(label="Utility cost per hour of bolted joints", value=0)
    depreciation_cost_bolted_joints = ft.TextField(label="Depreciation cost per hour of bolted joints", value=4.241)
    STba = ft.TextField(label="starting time bolted joints", value=0.02)

    #Economic Model Parameters
    cost_shipping = ft.TextField(label="Cost per kilometer of shipping", value=1.9625)
    cost_foundation = ft.TextField(label="Cost per squared meter of foundation", value=13.123)
    km_shipping = ft.TextField(label="Kilometers of shipping", value=100)
    kg_price = ft.TextField(label="Price of steel per kg", value=2)
    labor_cost_LAP = ft.TextField(label="Labor cost per hour of lift, alignment and positioning of heat exchanger", value=19.968)
    manipulation_time_assembly = ft.TextField(label="Manipulation Time per 1000kg of material", value=1)
    starting_time_LAP = ft.TextField(label="Starting time of lift, alignment and positioning of heat exchanger", value=0.5)
    k_scrap_factor_shell = ft.TextField(label="K scrap factor of shell", value=1)
    k_scrap_factor_tubes = ft.TextField(label="K scrap factor of tubes", value=0.9)
    k_scrap_factor_tubesheets = ft.TextField(label="K scrap factor of tubesheets", value=1)
    k_scrap_factor_heads = ft.TextField(label="K scrap factor of heads", value=1)
    k_scrap_factor_shell_nozzles = ft.TextField(label="K scrap factor of shell nozzles", value=1)
    k_scrap_factor_tubes_nozzles = ft.TextField(label="K scrap factor of tubes nozzles", value=1)
    k_scrap_factor_baffles = ft.TextField(label="K scrap factor of baffles", value=1)
    k_scrap_factor_flanges = ft.TextField(label="K scrap factor of flanges", value=1)
    k_scrap_factor_removable_covers = ft.TextField(label="K scrap factor of removable covers", value=1)
    k_scrap_factor_tie_rods = ft.TextField(label="K scrap factor of tie rods", value=0.8)
    k_scrap_factor_pass_partitions = ft.TextField(label="K scrap factor of pass partitions", value=1)
    k_scrap_factor_nozzles_flanges = ft.TextField(label="K scrap factor of flange nozzles", value=1)

    
    #Operational Settings

    #Plasma Cutting operation parameters
    labor_cost_plasma_cutting = ft.TextField(label="Labor cost per hour of plasma cutting", value=20)
    material_cost_plasma_cutting = ft.TextField(label="Material cost per hour of plasma cutting", value=9.94)
    utility_cost_plasma_cutting = ft.TextField(label="Utility cost per hour of plasma cutting", value=7.01)
    depreciation_cost_plasma_cutting = ft.TextField(label="Depreciation cost per hour of plasma cutting", value=4.24)
    starting_time_plasma_cutting = ft.TextField(label="starting time of plasma cutting", value=0.17)    
    manipulation_time_plasma_cutting = ft.TextField(label="manipulation time per 1000kg of material", value=1)
    spca = ft.TextField(label="spca", value=500)
    tpcs = ft.TextField(label="tpcs", value=25.4)
    operation_factor_plasma_cutting = ft.TextField(label="Operation factor of plasma cutting", value=0.5)
    ec = ft.TextField(label="Width of plasma cutting", value=3.25)


    #Drilling operation parameters
    labor_cost_drilling = ft.TextField(label="Labor cost per hour of drilling", value=20)
    material_cost_drilling = ft.TextField(label="Material cost per hour of drilling", value=6.71)
    utility_cost_drilling = ft.TextField(label="Utility cost per hour of drilling", value=0.19)
    depreciation_cost_drilling = ft.TextField(label="Depreciation cost per hour of drilling", value=4.24)
    starting_time_drilling = ft.TextField(label="Starting time of drilling", value=0.17)
    manipulation_time_drilling = ft.TextField(label="manipulation time per 1000kg of material", value=1)
    Stm = ft.TextField(label="Stm", value=2.5) #min
    rpm = ft.TextField(label="rpm of drilling tool", value=320)
    fnd = ft.TextField(label="Feed of tool", value=0.08)
    STdo = ft.TextField(label="Stdo", value=1)
    SDh = ft.TextField(label="Sdh", value=33.4)


    #Lathe operation parameters
    labor_cost_lathe = ft.TextField(label="Labor cost per hour of lathe", value=20)
    material_cost_lathe = ft.TextField(label="Material cost per hour of lathe", value=0.02)
    utility_cost_lathe = ft.TextField(label="Utility cost per hour of lathe", value=0.35)
    depreciation_cost_lathe = ft.TextField(label="Depreciation cost per hour of lathe", value=4.24)
    starting_time_lathe = ft.TextField(label="Starting time of lathe", value=1)
    manipulation_time_lathe = ft.TextField(label="manipulation time per 1000kg of material", value=1)
    operation_factor_lathe = ft.TextField(label="Operation factor of lathe", value=0.5)
    vcl = ft.TextField(label="vcl", value=200000)
    apl = ft.TextField(label="apl", value=2)
    fnl = ft.TextField(label="fnl", value=0.05)


    #Root pass welding operation parameters
    labor_cost_root_pass_welding = ft.TextField(label="Labor cost per hour of root pass welding", value=20)
    material_cost_root_pass_welding = ft.TextField(label="Material cost per hour of root pass welding", value=66.53)
    utility_cost_root_pass_welding = ft.TextField(label="Utility cost per hour of root pass welding", value=2.01)
    depreciation_cost_root_pass_welding = ft.TextField(label="Depreciation cost per hour of root pass welding", value=4.24)
    starting_time_root_pass_welding = ft.TextField(label="Starting time of root pass welding", value=0.17)
    manipulation_time_root_pass_welding = ft.TextField(label="manipulation time per 1000kg of material", value=1)
    density_filler_material = ft.TextField(label="Density of filler material for root pass", value=7850)
    nrp = ft.TextField(label="nrp", value=0.85)
    operation_factor_root_pass_welding = ft.TextField(label="Operation factor of root pass welding", value=0.5)
    rrp = ft.TextField(label="rrp", value=0.2)
    thickness_root_pass_welding = ft.TextField(label="Thickness of root pass welding", value=1.6)


    #Welding operation parameters
    labor_cost_welding = ft.TextField(label="Labor cost per hour of welding", value=20)
    material_cost_welding = ft.TextField(label="Material cost per hour of welding", value=66.53)
    utility_cost_welding = ft.TextField(label="Utility cost per hour of welding", value=2.01)
    depreciation_cost_welding = ft.TextField(label="Depreciation cost per hour of welding", value=4.24)
    starting_time_welding = ft.TextField(label="Starting time of welding", value=0.17)
    manipulation_time_welding = ft.TextField(label="manipulation time per 1000kg of material", value=1)
    density_filler_material = ft.TextField(label="Density of filler material for welding", value=7850)
    nw = ft.TextField(label="nw", value=0.85)
    operation_factor_welding = ft.TextField(label="Operation factor of welding", value=0.5)
    rw = ft.TextField(label="rw", value=0.2)
    thickness_root_pass_welding = ft.TextField(label="Thickness of root pass welding", value=1.6)
    

    #Blade Saw Cutting operation parameters
    labor_cost_blade_saw_cutting = ft.TextField(label="Labor cost per hour of blade saw cutting", value=20)
    material_cost_blade_saw_cutting = ft.TextField(label="Material cost per hour of blade saw cutting", value=0.01)
    utility_cost_blade_saw_cutting = ft.TextField(label="Utility cost per hour of blade saw cutting", value=0.1)
    depreciation_cost_blade_saw_cutting = ft.TextField(label="Depreciation cost per hour of blade saw cutting", value=4.24)
    starting_time_blade_saw_cutting = ft.TextField(label="Starting time of blade saw cutting", value=0.17)
    manipulation_time_blade_saw_cutting = ft.TextField(label="manipulation time per 1000kg of material", value=1)
    Crbs = ft.TextField(label="Crbs", value=700)
    operation_factor_blade_saw_cutting = ft.TextField(label="Operation factor of blade saw cutting", value=0.5)



    data_entry = ft.Text("Data Entry", size=24)
    sensit = ft.Text("Sensitivity", size=20, weight=ft.FontWeight.BOLD) #This shows the sensitivity of the change of cost FOB by changing a parameter


    def heat_map_make(e):
    
        ds = ft.TextField(label="Shell inside diameter", width=200, value=1219) #agregar unidades de todo
        ls = ft.TextField(label="Shell length", width=200, value=4876.8)
        dte = ft.TextField(label="Outside diameter of tubes", width=200, value=12.7)
        lay = ft.TextField(label="Pitch type", width=200, value="Triangle")
        ltp = ft.TextField(label="Tube pitch ratio", width=200, value=1.25)
        Nb  = ft.TextField(label="Number of baffles", width=200, value=14)
        BC  = ft.TextField(label="Baffle cut", width=200, value=25)
        number_passes = ft.TextField(label="Number of passes", width=200, value=4)
        Ntp = ft.TextField(label="Number of tubes per pass", width=200, value=449                                                                                                                                                                                       )
    
    #Parameters
        max_allowable_stress = ft.TextField(label="Maximum Allowable Stress", value=120)
        Sb  = ft.TextField(label="Maximum Allowable Bolt Stress", value=120)
        gasket_material_factor = ft.TextField(label="Gasket Material Factor", value=1.25)
        welding_efficiency = ft.TextField(label="Welding efficiency", value=0.8)
        corrosion_allowance = ft.TextField(label="Corrosion allowance", value=16)

    #Fluid Parameters
        shell_side_pressure = ft.TextField(label="Shell Side Pressure", value=0.2   )
        tube_side_pressure = ft.TextField(label="Tube Side pressure", value=0.2)
        shell_side_fluid_velocity = ft.TextField(label="Shell Side Fluid Velocoity", value=2)
        tube_side_fluid_velocity = ft.TextField(label="Tube Side Fluid Velocity", value=2)
        shell_side_mass_flow = ft.TextField(label="Shell Side Mass Flow", value=110)
        tube_side_mass_flow = ft.TextField(label="Tube Side Mass Flow", value=130)
        shell_side_fluid_density = ft.TextField(label="Shell Side Fluid Density", value=786)
        tube_side_fluid_density = ft.TextField(label="Tube Side Fluid Density", value=995)
  

    #Assembly Settings
        labor_cost_assembly = ft.TextField(label="Labor cost per hour of assembly", value=19.968)
        starting_time_assembly= ft.TextField(label="starting time assembly", value=0.02)
        manipulation_time_assembly = ft.TextField(label="Manipulation Time per 1000kg of material", value=1)
    
        labor_cost_bolted_joints = ft.TextField(label="Labor cost per hour of bolted joints", value=13)
        material_cost_bolted_joints = ft.TextField(label="Material cost per hour of bolted joints", value=0)
        utility_cost_bolted_joints = ft.TextField(label="Utility cost per hour of bolted joints", value=0)
        depreciation_cost_bolted_joints = ft.TextField(label="Depreciation cost per hour of bolted joints", value=4.241)
        STba = ft.TextField(label="starting time bolted joints", value=0.02)

    #Economic Model Parameters
        k_scrap_factor_shell = ft.TextField(label="K scrap factor of shell", value=1)
        k_scrap_factor_tubes = ft.TextField(label="K scrap factor of tubes", value=0.9)
        k_scrap_factor_tubesheets = ft.TextField(label="K scrap factor of tubesheets", value=1)
        k_scrap_factor_heads = ft.TextField(label="K scrap factor of heads", value=1)
        k_scrap_factor_shell_nozzles = ft.TextField(label="K scrap factor of shell nozzles", value=1)
        k_scrap_factor_tubes_nozzles = ft.TextField(label="K scrap factor of tubes nozzles", value=1)
        k_scrap_factor_baffles = ft.TextField(label="K scrap factor of baffles", value=1)
        k_scrap_factor_flanges = ft.TextField(label="K scrap factor of flanges", value=1)
        k_scrap_factor_removable_covers = ft.TextField(label="K scrap factor of removable covers", value=1)
        k_scrap_factor_tie_rods = ft.TextField(label="K scrap factor of tie rods", value=0.8)
        k_scrap_factor_pass_partitions = ft.TextField(label="K scrap factor of pass partitions", value=1)
        k_scrap_factor_nozzles_flanges = ft.TextField(label="K scrap factor of flange nozzles", value=1)

    
    #Operational Settings

    #Plasma Cutting operation parameters
        labor_cost_plasma_cutting = ft.TextField(label="Labor cost per hour of plasma cutting", value=20)
        material_cost_plasma_cutting = ft.TextField(label="Material cost per hour of plasma cutting", value=9.94)
        utility_cost_plasma_cutting = ft.TextField(label="Utility cost per hour of plasma cutting", value=7.01)
        depreciation_cost_plasma_cutting = ft.TextField(label="Depreciation cost per hour of plasma cutting", value=4.24)
        starting_time_plasma_cutting = ft.TextField(label="starting time of plasma cutting", value=0.17)    
        manipulation_time_plasma_cutting = ft.TextField(label="manipulation time per 1000kg of material", value=1)
        spca = ft.TextField(label="spca", value=500)
        tpcs = ft.TextField(label="tpcs", value=25.4)
        operation_factor_plasma_cutting = ft.TextField(label="Operation factor of plasma cutting", value=0.5)
        ec = ft.TextField(label="Width of plasma cutting", value=3.25)


    #Drilling operation parameters
        labor_cost_drilling = ft.TextField(label="Labor cost per hour of drilling", value=20)
        material_cost_drilling = ft.TextField(label="Material cost per hour of drilling", value=6.71)
        utility_cost_drilling = ft.TextField(label="Utility cost per hour of drilling", value=0.19)
        depreciation_cost_drilling = ft.TextField(label="Depreciation cost per hour of drilling", value=4.24)
        starting_time_drilling = ft.TextField(label="Starting time of drilling", value=0.17)
        manipulation_time_drilling = ft.TextField(label="manipulation time per 1000kg of material", value=1)
        Stm = ft.TextField(label="Stm", value=2.5) #min
        rpm = ft.TextField(label="rpm of drilling tool", value=320)
        fnd = ft.TextField(label="Feed of tool", value=0.08)
        STdo = ft.TextField(label="Stdo", value=1)
        SDh = ft.TextField(label="Sdh", value=33.4)


    #Lathe operation parameters
        labor_cost_lathe = ft.TextField(label="Labor cost per hour of lathe", value=20)
        material_cost_lathe = ft.TextField(label="Material cost per hour of lathe", value=0.02)
        utility_cost_lathe = ft.TextField(label="Utility cost per hour of lathe", value=0.35)
        depreciation_cost_lathe = ft.TextField(label="Depreciation cost per hour of lathe", value=4.24)
        starting_time_lathe = ft.TextField(label="Starting time of lathe", value=1)
        manipulation_time_lathe = ft.TextField(label="manipulation time per 1000kg of material", value=1)
        operation_factor_lathe = ft.TextField(label="Operation factor of lathe", value=0.5)
        vcl = ft.TextField(label="vcl", value=200000)
        apl = ft.TextField(label="apl", value=2)
        fnl = ft.TextField(label="fnl", value=0.05)


    #Root pass welding operation parameters
        labor_cost_root_pass_welding = ft.TextField(label="Labor cost per hour of root pass welding", value=20)
        material_cost_root_pass_welding = ft.TextField(label="Material cost per hour of root pass welding", value=66.53)
        utility_cost_root_pass_welding = ft.TextField(label="Utility cost per hour of root pass welding", value=2.01)
        depreciation_cost_root_pass_welding = ft.TextField(label="Depreciation cost per hour of root pass welding", value=4.24)
        starting_time_root_pass_welding = ft.TextField(label="Starting time of root pass welding", value=0.17)
        manipulation_time_root_pass_welding = ft.TextField(label="manipulation time per 1000kg of material", value=1)
        density_filler_material = ft.TextField(label="Density of filler material for root pass", value=7850)
        nrp = ft.TextField(label="nrp", value=0.85)
        operation_factor_root_pass_welding = ft.TextField(label="Operation factor of root pass welding", value=0.5)
        rrp = ft.TextField(label="rrp", value=0.2)
        thickness_root_pass_welding = ft.TextField(label="Thickness of root pass welding", value=1.6)


    #Welding operation parameters
        labor_cost_welding = ft.TextField(label="Labor cost per hour of welding", value=20)
        material_cost_welding = ft.TextField(label="Material cost per hour of welding", value=66.53)
        utility_cost_welding = ft.TextField(label="Utility cost per hour of welding", value=2.01)
        depreciation_cost_welding = ft.TextField(label="Depreciation cost per hour of welding", value=4.24)
        starting_time_welding = ft.TextField(label="Starting time of welding", value=0.17)
        manipulation_time_welding = ft.TextField(label="manipulation time per 1000kg of material", value=1)
        density_filler_material = ft.TextField(label="Density of filler material for welding", value=7850)
        nw = ft.TextField(label="nw", value=0.85)
        operation_factor_welding = ft.TextField(label="Operation factor of welding", value=0.5)
        rw = ft.TextField(label="rw", value=0.2)
        thickness_root_pass_welding = ft.TextField(label="Thickness of root pass welding", value=1.6)
    

    #Blade Saw Cutting operation parameters
        labor_cost_blade_saw_cutting = ft.TextField(label="Labor cost per hour of blade saw cutting", value=20)
        material_cost_blade_saw_cutting = ft.TextField(label="Material cost per hour of blade saw cutting", value=0.01)
        utility_cost_blade_saw_cutting = ft.TextField(label="Utility cost per hour of blade saw cutting", value=0.1)
        depreciation_cost_blade_saw_cutting = ft.TextField(label="Depreciation cost per hour of blade saw cutting", value=4.24)
        starting_time_blade_saw_cutting = ft.TextField(label="Starting time of blade saw cutting", value=0.17)
        manipulation_time_blade_saw_cutting = ft.TextField(label="manipulation time per 1000kg of material", value=1)
        Crbs = ft.TextField(label="Crbs", value=700)
        operation_factor_blade_saw_cutting = ft.TextField(label="Operation factor of blade saw cutting", value=0.5)

        increased_parameters = ["Labor cost assembly", "Labor cost bolted joints", "Material cost bolted joints", "Utility cost bolted joints", "Depreciation cost bolted joints","Labor cost Plasma cutting", "Material cost Plasma cutting", "Utility cost Plasma cutting", "Depreciation cost Plasma cutting","Labor cost drilling", "Material cost drilling", "Utility cost drilling", "Depreciation cost drilling", "Labor cost lathe", "Material cost lathe", "Utility cost lathe", "Depreciation cost lathe","Labor cost root pass welding", "Material cost root pass welding", "Utility cost root pass welding", "Depreciation cost root pass welding", "Labor cost welding", "Material cost welding", "Utility cost welding", "Depreciation cost welding", "Labor cost blade saw cutting", "Material cost blade saw cutting", "Utility cost blade saw cutting", "Depreciation cost blade saw cutting"]

        name_examples = ["Example 1","Example 2","Example 3","Example 4","Example 5","Example 6","Example 7","Example 8","Example 9","Example 10","Example 11","Example 12","Example 13","Example 14","Example 15","Example 16","Example 17","Example 18","Example 19","Example 20","Example 21","Example 22","Example 23","Example 24","Example 25","Example 26","Example 27","Example 28","Example 29","Example 30","Example 31","Example 32","Example 33","Example 34"]
        
        
        example_1={"ds":1219.2,"ls":4876.8,"Nb":14,"dte":12.7,"ltp":1.25,"lay":"Triangle","number_passes":4, "Ntp":449, "shell_side_mass_flow":110,"tube_side_mass_flow":130,"shell_side_fluid_density":786,  "tube_side_fluid_density":995}
        example_2={"ds":787.4,"ls":6096,"Nb":17,"dte":12.7,"ltp":1.25,"lay":"Square","number_passes":2,"Ntp":376,"shell_side_mass_flow":50,"tube_side_mass_flow":130,"shell_side_fluid_density":786, "tube_side_fluid_density":995}

        example_3={"ds":838.2,"ls":4876.8,"Nb":14,"dte":12.7,"ltp":1.5,"lay":"Square","number_passes":6,"Ntp":90,"shell_side_mass_flow":56.6,"tube_side_mass_flow":27.8,"shell_side_fluid_density":995,"tube_side_fluid_density":750}

        example_4={"ds":1524, "ls":6096, "Nb":5,"dte":12.7,"ltp":1.5, "lay":"Square", "number_passes":6,"Ntp":325, "shell_side_mass_flow":120, "tube_side_mass_flow":69.4,"tube_side_fluid_density":750,"shell_side_fluid_density":995}

        example_5={"ds":838.2,"ls":3657.6,"Nb":8,"dte":12.7,"ltp":1.5,"lay":"Square","number_passes":4, "Ntp":140, "shell_side_mass_flow":100,"tube_side_mass_flow":40, "shell_side_fluid_density":750, "tube_side_fluid_density":888}

        example_6={"ds":787.4, "ls":6096, "Nb":16,"dte":12.7, "ltp":1.25,"lay":"Square", "number_passes":1, "Ntp":777,"shell_side_mass_flow":55.6, "tube_side_mass_flow":130, "shell_side_fluid_density":789, "tube_side_fluid_density":995}



        example_7={"ds":838.2, "ls":3657.6, "Nb":7,"dte":12.7, "ltp":1.25,"lay":"Square", "number_passes":6, "Ntp":130, "shell_side_mass_flow":133.33, "tube_side_mass_flow":40, "shell_side_fluid_density":1080, "tube_side_fluid_density":888}
        
        example_8={"ds":1524,"ls":4876.8,"Nb":15,"dte":12.7,"ltp":1.25,"lay":"Square","number_passes":4,"Ntp":720,"shell_side_mass_flow":83.3,"tube_side_mass_flow":120,"shell_side_fluid_density":1080,"tube_side_fluid_density":995}

        example_9={"ds":787.4,"ls":4876.8,"Nb":10,"dte":12.7,"ltp":1.25,"lay":"Square","number_passes":2,"Ntp":376,"shell_side_mass_flow":80,"tube_side_mass_flow":100,"shell_side_fluid_density":789,"tube_side_fluid_density":736}
        example_10={"ds":990.6,"ls":3048,"Nb":3,"dte":12.7,"ltp":1.25,"lay":"Triangle","number_passes":6,"Ntp":217,"shell_side_mass_flow":100,"tube_side_mass_flow":80,"shell_side_fluid_density":736,"tube_side_fluid_density":789}

        example_11={"ds":1371.6,"ls":6096,"Nb":7,"dte":12.7,"ltp":1.33,"lay":"Square","number_passes":2,"Ntp":1058,"shell_side_mass_flow":110,"tube_side_mass_flow":120,"shell_side_fluid_density":786,"tube_side_fluid_density":995}

        example_12={"ds":939.8,"ls":4876.8,"Nb":12,"dte":12.7,"ltp":1.25,"lay":"Square","number_passes":2,"Ntp":544,"shell_side_mass_flow":50,"tube_side_mass_flow":130,"shell_side_fluid_density":786,"tube_side_fluid_density":995}

        example_13={"ds":889,"ls":4876.8,"Nb":7,"dte":19.05,"ltp":1.25,"lay":"Square","number_passes":6,"Ntp":82,"shell_side_mass_flow":56.6,"tube_side_mass_flow":27.8,"shell_side_fluid_density":995,"tube_side_fluid_density":750}

        example_14={"ds":1524,"ls":6096,"Nb":4,"dte":12.7,"ltp":1.5,"lay":"Square","number_passes":6,"Ntp":375,"shell_side_mass_flow":120,"tube_side_mass_flow":69.4,"shell_side_fluid_density":995,"tube_side_fluid_density":750}

        example_15={"ds":990.6,"ls":3657.6,"Nb":3,"dte":12.7,"ltp":1.5,"lay":"Square","number_passes":4,"Ntp":201,"shell_side_mass_flow":100,"tube_side_mass_flow":40,"shell_side_fluid_density":750,"tube_side_fluid_density":888}

        example_16={"ds":1066.8,"ls":6096,"Nb":14,"dte":19.05,"ltp":1.33,"lay":"Square","number_passes":1,"Ntp":716,"shell_side_mass_flow":55.6,"tube_side_mass_flow":130,"shell_side_fluid_density":789,"tube_side_fluid_density":995}

        example_17={"ds":990.6,"ls":3657.6,"Nb":3,"dte":12.7,"ltp":1.33,"lay":"Square","number_passes":6,"Ntp":166,"shell_side_mass_flow":133,"tube_side_mass_flow":40,"shell_side_fluid_density":1080,"tube_side_fluid_density":888}
        example_18={"ds":1524,"ls":4876.8,"Nb":14,"dte":12.7,"ltp":1.25,"lay":"Square","number_passes":2,"Ntp":1489,"shell_side_mass_flow":83.3,"tube_side_mass_flow":120,"shell_side_fluid_density":1080,"tube_side_fluid_density":995}


        example_19={"ds":1219.2,"ls":6096,"Nb":5,"dte":25.4,"ltp":1.25,"lay":"Square","number_passes":2,"Ntp":330,"shell_side_mass_flow":60,"tube_side_mass_flow":100,"shell_side_fluid_density":789,"tube_side_fluid_density":736}




        example_20={"ds":1371.6,"ls":4876.8,"Nb":3,"dte":25.4,"ltp":1.33,"lay":"Triangle","number_passes":4,"Ntp":207,"shell_side_mass_flow":100,"tube_side_mass_flow":80,"shell_side_fluid_density":736,"tube_side_fluid_density":789}



        example_21={"ds":838.2,"ls":5029.2,"Nb":8,"dte":12.7,"ltp":1.5,"lay":"Triangle","number_passes":4,"Ntp":140,"shell_side_mass_flow":100,"tube_side_mass_flow":40,"shell_side_fluid_density":750,"tube_side_fluid_density":888}

        
        example_22={"ds":838.2,"ls":6705.6,"Nb":8,"dte":12.7,"ltp":1.5,"lay":"Triangle","number_passes":4,"Ntp":140,"shell_side_mass_flow":100,"tube_side_mass_flow":40,"shell_side_fluid_density":750,"tube_side_fluid_density":888}




        example_23={"ds":889,"ls":7112,"Nb":7,"dte":19.05,"ltp":1.25,"lay":"Square","number_passes":6,"Ntp":82,"shell_side_mass_flow":56.6,"tube_side_mass_flow":27.8,"shell_side_fluid_density":995,"tube_side_fluid_density":750}



        example_24={"ds":1270,"ls":7620,"Nb":7,"dte":19.05, "ltp":1.25,"lay":"Square","number_passes":6,"Ntp":82,"shell_side_mass_flow":56.6,"tube_side_mass_flow":27.8,"shell_side_fluid_density":995,"tube_side_fluid_density":750}




        example_25={"ds":1066.8,"ls":8534.4,"Nb":4,"dte":19.05,"ltp":1.33,"lay":"Square","number_passes":1,"Ntp":716,"shell_side_mass_flow":56.6,"tube_side_mass_flow":130,"shell_side_fluid_density":789,"tube_side_fluid_density":995}



        example_26={"ds":1270,"ls":9601.2,"Nb":14,"dte":19.05,"ltp":1.33,"lay":"Square","number_passes":1,"Ntp":716,"shell_side_mass_flow":56.6,"tube_side_mass_flow":130,"shell_side_fluid_density":789,"tube_side_fluid_density":995}



        example_27={"ds":1371.6,"ls":8229.6,"Nb":3,"dte":25.4,"ltp":1.33,"lay":"Triangle","number_passes":4,"Ntp":207,"shell_side_mass_flow":100,"tube_side_mass_flow":80,"shell_side_fluid_density":736,"tube_side_fluid_density":789}


        example_28={"ds":990.6,"ls":5943.6,"Nb":3,"dte":25.4,"ltp":1.33,"lay":"Triangle","number_passes":4,"Ntp":207,"shell_side_mass_flow":100,"tube_side_mass_flow":80,"shell_side_fluid_density":736,"tube_side_fluid_density":789}

        example_29={"ds":990.6,"ls":6934.2,"Nb":3,"dte":12.7,"ltp":1.33,"lay":"Square","number_passes":6,"Ntp":166,"shell_side_mass_flow":133.33,"tube_side_mass_flow":40,"shell_side_fluid_density":1080,"tube_side_fluid_density":888}


        example_30={"ds":1066.8,"ls":7467.6,"Nb":14,"dte":19.05,"ltp":1.33,"lay":"Square","number_passes":1,"Ntp":716,"shell_side_mass_flow":55.6,"tube_side_mass_flow":130,"shell_side_fluid_density":789,"tube_side_fluid_density":995}

        example_31={"ds":1066.8,"ls":9601.2,"Nb":14,"dte":19.05,"ltp":1.33,"lay":"Square","number_passes":1,"Ntp":716,"shell_side_mass_flow":55.6,"tube_side_mass_flow":130,"shell_side_fluid_density":789,"tube_side_fluid_density":995}

        example_32={"ds":1219.2,"ls":7315.2,"Nb":14,"dte":12.7, "ltp":1.25,"lay":"Square","number_passes":4,"Ntp":449,"shell_side_mass_flow":110,"tube_side_mass_flow":130,"shell_side_fluid_density":789,"tube_side_fluid_density":995}

        example_33={"ds":1016,"ls":6934.2,"Nb":3,"dte":12.7,"ltp":1.25,"lay":"Square","number_passes":6,"Ntp":217,"shell_side_mass_flow":100,"tube_side_mass_flow":80,"shell_side_fluid_density":736,"tube_side_fluid_density":789}

        example_34={"ds":228.6,"ls":2311.4,"Nb":17,"dte":3.175,"ltp":1.25,"lay":"Triangle","number_passes":4,"Ntp":64,"shell_side_mass_flow":5,"tube_side_mass_flow":5,"shell_side_fluid_density":786.4,"tube_side_fluid_density":995}

        #cost_raw_material_shell = total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[2]

        #cost_raw_material_tubesheets =total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[3]
 
        #cost_raw_material_tubes=total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[4]

        #cost_raw_material_heads=total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[5]

        #cost_raw_material_nozzles=total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[6]

        #cost_raw_material_flanges=total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[7]

        #cost_raw_material_removable_covers=total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[8]

        #cost_raw_material_pass_partitions=total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[9]

        #cost_raw_material_tie_rods=total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[10]

        #cost_raw_material_nozzles_flanges=total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[11]
        #cost_raw_material_baffles =total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[12]

        examples=[example_1,example_2,example_3,example_4,example_5,example_6,example_7,example_8,example_9,example_10,example_11,example_12,example_13,example_14,example_15,example_16,example_17,example_18,example_19,example_20,example_21,example_22,example_23,example_24,example_25,example_26,example_27,example_28,example_29,example_30,example_31,example_32,example_33,example_34]

        parameters=[labor_cost_assembly,labor_cost_bolted_joints,material_cost_bolted_joints,utility_cost_bolted_joints,depreciation_cost_bolted_joints,labor_cost_plasma_cutting,material_cost_plasma_cutting,utility_cost_plasma_cutting,depreciation_cost_plasma_cutting,labor_cost_drilling,material_cost_drilling,utility_cost_drilling,depreciation_cost_drilling,labor_cost_lathe,material_cost_lathe,utility_cost_lathe,depreciation_cost_lathe,labor_cost_root_pass_welding,material_cost_root_pass_welding,utility_cost_root_pass_welding,depreciation_cost_root_pass_welding,labor_cost_welding,material_cost_welding,utility_cost_welding,depreciation_cost_welding,labor_cost_blade_saw_cutting,material_cost_blade_saw_cutting,utility_cost_blade_saw_cutting,depreciation_cost_blade_saw_cutting]


        """plates_table = [
                {"Width": 3000,"Length": 10000,"Thickness":    4.7625,"Price": 2803.92},
                {"Width": 3000,"Length": 10000,"Thickness":    6.35,"Price": 3738.56},
                {"Width": 3000,"Length": 10000,"Thickness":    9.525,"Price":5607.84},
                {"Width": 3000,"Length": 10000,"Thickness":    12.7,"Price":7477.13},
                {"Width": 3000,"Length": 10000,"Thickness":15.875,"Price": 9346.41},
                {"Width": 3000,"Length": 10000,"Thickness":     19.05,"Price": 11215.69},
                {"Width": 3000,"Length": 10000,"Thickness":22.225,"Price": 13084.97},
                {"Width": 3000,"Length": 10000,"Thickness":25.4,"Price": 14954.25},
                {"Width": 3000,"Length": 10000,"Thickness":38.1,"Price": 22431.38},
                {"Width": 3000,"Length": 10000,"Thickness":44.45,"Price": 26169.94},
                {"Width": 3000,"Length": 10000,"Thickness":50.8,"Price":29908.5},
                {"Width": 3000,"Length": 10000,"Thickness":63.5,"Price":37385.63},
                {"Width": 3000,"Length": 10000,"Thickness":69.85,"Price": 41124.19},
                {"Width": 3000,"Length": 10000,"Thickness":76.2,"Price":44862.75},
                {"Width": 3000,"Length": 10000,"Thickness":101.6,"Price":59817}
                ]

#t
#Rods
    #OD    Lr   $
            rods_table = [
                {"Outside diameter": 3.175,"Length": 10000,"Price": 1.55},
                {"Outside diameter": 6.35,"Length": 10000,"Price": 6.31},
                {"Outside diameter": 9.525,"Length":10000,"Price": 14},
                {"Outside diameter": 12.7,"Length": 10000,"Price": 24.86},
                {"Outside diameter": 15.875,"Length": 10000,"Price":38.87},
                {"Outside diameter": 19.05,"Length": 10000,"Price": 55.94},
                {"Outside diameter": 22.225,"Length": 10000,"Price": 76.17},
                {"Outside diameter": 25.4,"Length": 10000,"Price": 99.44},
                {"Outside diameter": 28.575,"Length": 10000,"Price": 125.9},
                {"Outside diameter": 31.75,"Length": 10000,"Price":155.38},
                {"Outside diameter": 34.925,"Length": 10000,"Price": 188.06},
                {"Outside diameter": 38.1,"Length": 10000,"Price": 223.74},
                {"Outside diameter": 41.275,"Length": 10000,"Price": 262.91}
                ]

#Tubes from SCHEDULE 40s and 40 https://www.trupply.com/products/seamless-pipe-a106b?variant=41891233464482
   #dte     L   t       $
            schedule40_table = [
                {"Outside diameter": 10.3,"Length":10000,"Wall thickness": 1.73,"Price": 9.14},
                {"Outside diameter": 13.7,"Length":10000,"Wall thickness": 2.24,"Price":15.83},
                {"Outside diameter": 17.1,"Length":10000,"Wall thickness": 2.31,"Price":21.06},
                {"Outside diameter": 21.30,"Length":10000,"Wall thickness": 2.77,"Price":31.65},
                {"Outside diameter": 26.7,"Length":10000,"Wall thickness": 2.87,"Price":42.17},
                {"Outside diameter": 33.4,"Length":10000,"Wall thickness": 3.38,"Price":62.56},
                {"Outside diameter": 42.2,"Length":10000,"Wall thickness": 3.56,"Price":84.81},
                {"Outside diameter": 48.3,"Length":10000,"Wall thickness": 3.68,"Price":101.24},
                {"Outside diameter": 60.3,"Length":10000,"Wall thickness": 3.91,"Price":135.94},
                {"Outside diameter": 73,"Length":10000,"Wall thickness": 5.16,"Price":215.82},
                {"Outside diameter": 88.9,"Length":10000,"Wall thickness": 5.49,"Price":282.33},
                {"Outside diameter": 101.6,"Length":10000,"Wall thickness": 5.74,"Price": 339.24},
                {"Outside diameter": 114.3,"Length":10000,"Wall thickness": 6.02,"Price": 401.89},
                {"Outside diameter": 141.3,"Length":10000,"Wall thickness": 6.55,"Price": 544.16},
                {"Outside diameter": 168.3,"Length":10000,"Wall thickness": 7.11,"Price":706.59},
                {"Outside diameter": 219.1,"Length":10000,"Wall thickness": 8.18,"Price":1063.73},
                {"Outside diameter": 273.0,"Length":10000,"Wall thickness": 9.27,"Price":1507.3},
                {"Outside diameter": 323.8,"Length":10000,"Wall thickness": 10.31,"Price": 1992.7},
                {"Outside diameter": 355.6,"Length":10000,"Wall thickness": 11.13,"Price":2363.77},
                {"Outside diameter": 406.4,"Length":10000,"Wall thickness": 12.7,"Price":3082.68},
                {"Outside diameter": 457.0,"Length":10000,"Wall thickness": 14.27,"Price":3895.13},
                {"Outside diameter": 508.0,"Length":10000,"Wall thickness": 15.09,"Price":4585.81},
                {"Outside diameter": 610.0,"Length":10000,"Wall thickness": 17.48,"Price":63858.63}
                ]"""


        heat_map2=[]
        raw_material_parameters = ["Price of plates","Price of tubes","Price of rods"]

        heat_map=[]
        for example in examples:
            plates_table = [
{"Width": 3000,"Length": 10000,"Thickness":    4.7625,"Price":2243.14},
{"Width": 3000,"Length": 10000,"Thickness":    6.35,"Price":    2990.85},
{"Width": 3000,"Length": 10000,"Thickness":    9.525,"Price": 4468.28},
{"Width": 3000,"Length": 10000,"Thickness":    12.7,"Price":  5981.7},
{"Width": 3000,"Length": 10000,"Thickness":15.875,"Price":    7477.15},
{"Width": 3000,"Length": 10000,"Thickness":     19.05,"Price":   8972.55},
{"Width": 3000,"Length": 10000,"Thickness":22.225,"Price": 10467.98},
{"Width": 3000,"Length": 10000,"Thickness":25.4,"Price":   11963.4},
{"Width": 3000,"Length": 10000,"Thickness":38.1,"Price": 17945.1 },
{"Width": 3000,"Length": 10000,"Thickness":44.45,"Price":20935.95},
{"Width": 3000,"Length": 10000,"Thickness":50.8,"Price":23926.8},
{"Width": 3000,"Length": 10000,"Thickness":63.5,"Price":29908.5},
{"Width": 3000,"Length": 10000,"Thickness":69.85,"Price":32899.35},
{"Width": 3000,"Length": 10000,"Thickness":76.2,"Price":35890.2},
{"Width": 3000,"Length": 10000,"Thickness":101.6,"Price":47853.6}
]

#t
#Rods
    #OD    Lr   $
            rods_table = [
{"Outside diameter": 3.175,"Length": 10000,"Price": 1.24},
{"Outside diameter": 6.35,"Length": 10000,"Price": 5.05},
{"Outside diameter": 9.525,"Length":10000,"Price": 11.20},
{"Outside diameter": 12.7,"Length": 10000,"Price":19.89},
{"Outside diameter": 15.875,"Length": 10000,"Price": 31.09},
{"Outside diameter": 19.05,"Length": 10000,"Price": 44.75},
{"Outside diameter": 22.225,"Length": 10000,"Price": 60.94},
{"Outside diameter": 25.4,"Length": 10000,"Price": 79.55},
{"Outside diameter": 28.575,"Length": 10000,"Price": 100.72},
{"Outside diameter": 31.75,"Length": 10000,"Price":124.3},
{"Outside diameter": 34.925,"Length": 10000,"Price": 150.45},
{"Outside diameter": 38.1,"Length": 10000,"Price":178.99},
{"Outside diameter": 41.275,"Length": 10000,"Price": 210.32}
]

#Tubes from SCHEDULE 40s and 40 https://www.trupply.com/products/seamless-pipe-a106b?variant=41891233464482
   #dte     L   t       $
            schedule40_table = [
{"Outside diameter": 10.3,"Length":10000,"Wall thickness": 1.73,"Price": 7.31},
{"Outside diameter": 13.7,"Length":10000,"Wall thickness": 2.24,"Price": 12.66},
{"Outside diameter": 17.1,"Length":10000,"Wall thickness": 2.31,"Price": 16.85},
{"Outside diameter": 21.30,"Length":10000,"Wall thickness": 2.77,"Price": 25.32},
{"Outside diameter": 26.7,"Length":10000,"Wall thickness": 2.87,"Price":33.73},
{"Outside diameter": 33.4,"Length":10000,"Wall thickness": 3.38,"Price": 50.05},
{"Outside diameter": 42.2,"Length":10000,"Wall thickness": 3.56,"Price": 67.85},
{"Outside diameter": 48.3,"Length":10000,"Wall thickness": 3.68,"Price": 80.99},
{"Outside diameter": 60.3,"Length":10000,"Wall thickness": 3.91,"Price": 108.75},
{"Outside diameter": 73,"Length":10000,"Wall thickness": 5.16,"Price": 172.66},
{"Outside diameter": 88.9,"Length":10000,"Wall thickness": 5.49,"Price": 225.86},
{"Outside diameter": 101.6,"Length":10000,"Wall thickness": 5.74,"Price": 271.39},
{"Outside diameter": 114.3,"Length":10000,"Wall thickness": 6.02,"Price": 321.51},
{"Outside diameter": 141.3,"Length":10000,"Wall thickness": 6.55,"Price": 435.33},
{"Outside diameter": 168.3,"Length":10000,"Wall thickness": 7.11,"Price": 565.27},
{"Outside diameter": 219.1,"Length":10000,"Wall thickness": 8.18,"Price": 850.98},
{"Outside diameter": 273.0,"Length":10000,"Wall thickness": 9.27,"Price": 1205.84},
{"Outside diameter": 323.8,"Length":10000,"Wall thickness": 10.31,"Price": 1594.16},
{"Outside diameter": 355.6,"Length":10000,"Wall thickness": 11.13,"Price": 1891.02},
{"Outside diameter": 406.4,"Length":10000,"Wall thickness": 12.7,"Price": 2466.15},
{"Outside diameter": 457.0,"Length":10000,"Wall thickness": 14.27,"Price": 3116.11},
{"Outside diameter": 508.0,"Length":10000,"Wall thickness": 15.09,"Price": 3668.65},
{"Outside diameter": 610.0,"Length":10000,"Wall thickness": 17.48,"Price": 5108.51}
]

            ds.value=example["ds"]
            ls.value=example["ls"]
            Nb.value=example["Nb"]
            dte.value=example["dte"]

            number_passes.value=example["number_passes"]
            Ntp.value=example["Ntp"]
            lay.value=example["lay"]
            ltp.value = example["ltp"]
            shell_side_mass_flow.value=example["shell_side_mass_flow"]
            tube_side_mass_flow.value=example["tube_side_mass_flow"]
            shell_side_fluid_density.value = example["shell_side_fluid_density"]
            tube_side_fluid_density.value = example["tube_side_fluid_density"]
            sensitivities=[]
            sensitivities2=[]
            for parameter in parameters:
                sensitivity_amount = 0.01
                if parameter.value==0:
                    h=1
                else:
                    h = parameter.value * sensitivity_amount #amount of change in parameter, sensityvity_amount shoulde be 0.25 or 0.5
#from here on, calculates onlye sensitivity parameter
                original_value_parameter = parameter.value
                total_cost_heat_exchanger_FOB_before = total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]
                factor_p_over_y = parameter.value/total_cost_heat_exchanger_FOB_before

                print(f"Factor p over y is {factor_p_over_y}")
                parameter.value += h  #increase the parameter in an amount h
                total_cost_heat_exchanger_FOB_after = total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]
                sensitivity = factor_p_over_y * ((total_cost_heat_exchanger_FOB_after-total_cost_heat_exchanger_FOB_before)/h)
                sensitivity = round(sensitivity,4)
#up to h, sensitivity is definded
                parameter.value= original_value_parameter #parameter goes back to original value        
                sensitivities.append(sensitivity)

            heat_map.append(sensitivities)
            print(f"Heat map is: \n\n\n {heat_map}")


            wb = Workbook()
            ws = wb.active
            ws.title = "Heat Map 1"
            for row_index, row_data in enumerate(heat_map):
                for col_index, cell_value in enumerate(row_data):
                    ws.cell(row=row_index+1, column=col_index+1, value=cell_value)

            wb.save("Heat map 1.xlsx")


################# Parameters calculations ends here, below this, we calculate sensitivity for raw material##############
#we take the average of all plates

            total_cost_heat_exchanger_FOB_before_plates = total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]

            plates_table = [
            {"Width": 3000,"Length": 10000,"Thickness":    4.7625,"Price": 2803.92},
            {"Width": 3000,"Length": 10000,"Thickness":    6.35,"Price": 3738.56},
            {"Width": 3000,"Length": 10000,"Thickness":    9.525,"Price":5607.84},
            {"Width": 3000,"Length": 10000,"Thickness":    12.7,"Price":7477.13},
            {"Width": 3000,"Length": 10000,"Thickness":15.875,"Price": 9346.41},
            {"Width": 3000,"Length": 10000,"Thickness":     19.05,"Price": 11215.69},
            {"Width": 3000,"Length": 10000,"Thickness":22.225,"Price": 13084.97},
            {"Width": 3000,"Length": 10000,"Thickness":25.4,"Price": 14954.25},
            {"Width": 3000,"Length": 10000,"Thickness":38.1,"Price": 22431.38},
            {"Width": 3000,"Length": 10000,"Thickness":44.45,"Price": 26169.94},
            {"Width": 3000,"Length": 10000,"Thickness":50.8,"Price":29908.5},
            {"Width": 3000,"Length": 10000,"Thickness":63.5,"Price":37385.63},
            {"Width": 3000,"Length": 10000,"Thickness":69.85,"Price": 41124.19},
            {"Width": 3000,"Length": 10000,"Thickness":76.2,"Price":44862.75},
            {"Width": 3000,"Length": 10000,"Thickness":101.6,"Price":59817}
            ]

            total_cost_heat_exchanger_FOB_after_plates = total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]
            sensitivity_plates = 4 * ((total_cost_heat_exchanger_FOB_after_plates-total_cost_heat_exchanger_FOB_before_plates)/total_cost_heat_exchanger_FOB_before_plates)
            sensitivity_plates = round(sensitivity_plates, 4)
                #up to h, sensitivity is definded

            sensitivities2.append(sensitivity_plates)
            plates_table = [
{"Width": 3000,"Length": 10000,"Thickness":    4.7625,"Price":2243.14},
{"Width": 3000,"Length": 10000,"Thickness":    6.35,"Price":    2990.85},
{"Width": 3000,"Length": 10000,"Thickness":    9.525,"Price": 4468.28},
{"Width": 3000,"Length": 10000,"Thickness":    12.7,"Price":  5981.7},
{"Width": 3000,"Length": 10000,"Thickness":15.875,"Price":    7477.15},
{"Width": 3000,"Length": 10000,"Thickness":     19.05,"Price":   8972.55},
{"Width": 3000,"Length": 10000,"Thickness":22.225,"Price": 10467.98},
{"Width": 3000,"Length": 10000,"Thickness":25.4,"Price":   11963.4},
{"Width": 3000,"Length": 10000,"Thickness":38.1,"Price": 17945.1 },
{"Width": 3000,"Length": 10000,"Thickness":44.45,"Price":20935.95},
{"Width": 3000,"Length": 10000,"Thickness":50.8,"Price":23926.8},
{"Width": 3000,"Length": 10000,"Thickness":63.5,"Price":29908.5},
{"Width": 3000,"Length": 10000,"Thickness":69.85,"Price":32899.35},
{"Width": 3000,"Length": 10000,"Thickness":76.2,"Price":35890.2},
{"Width": 3000,"Length": 10000,"Thickness":101.6,"Price":47853.6}
]



            total_cost_heat_exchanger_FOB_before_tubes =total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]


            schedule40_table = [
            {"Outside diameter": 10.3,"Length":10000,"Wall thickness": 1.73,"Price": 9.14},
            {"Outside diameter": 13.7,"Length":10000,"Wall thickness": 2.24,"Price":15.83},
            {"Outside diameter": 17.1,"Length":10000,"Wall thickness": 2.31,"Price":21.06},
            {"Outside diameter": 21.30,"Length":10000,"Wall thickness": 2.77,"Price":31.65},
            {"Outside diameter": 26.7,"Length":10000,"Wall thickness": 2.87,"Price":42.17},
            {"Outside diameter": 33.4,"Length":10000,"Wall thickness": 3.38,"Price":62.56},
            {"Outside diameter": 42.2,"Length":10000,"Wall thickness": 3.56,"Price":84.81},
            {"Outside diameter": 48.3,"Length":10000,"Wall thickness": 3.68,"Price":101.24},
            {"Outside diameter": 60.3,"Length":10000,"Wall thickness": 3.91,"Price":135.94},
            {"Outside diameter": 73,"Length":10000,"Wall thickness": 5.16,"Price":215.82},
            {"Outside diameter": 88.9,"Length":10000,"Wall thickness": 5.49,"Price":282.33},
            {"Outside diameter": 101.6,"Length":10000,"Wall thickness": 5.74,"Price": 339.24},
            {"Outside diameter": 114.3,"Length":10000,"Wall thickness": 6.02,"Price": 401.89},
            {"Outside diameter": 141.3,"Length":10000,"Wall thickness": 6.55,"Price": 544.16},
            {"Outside diameter": 168.3,"Length":10000,"Wall thickness": 7.11,"Price":706.59},
            {"Outside diameter": 219.1,"Length":10000,"Wall thickness": 8.18,"Price":1063.73},
            {"Outside diameter": 273.0,"Length":10000,"Wall thickness": 9.27,"Price":1507.3},
            {"Outside diameter": 323.8,"Length":10000,"Wall thickness": 10.31,"Price": 1992.7},
            {"Outside diameter": 355.6,"Length":10000,"Wall thickness": 11.13,"Price":2363.77},
            {"Outside diameter": 406.4,"Length":10000,"Wall thickness": 12.7,"Price":3082.68},
            {"Outside diameter": 457.0,"Length":10000,"Wall thickness": 14.27,"Price":3895.13},
            {"Outside diameter": 508.0,"Length":10000,"Wall thickness": 15.09,"Price":4585.81},
            {"Outside diameter": 610.0,"Length":10000,"Wall thickness": 17.48,"Price":6385.63}
            ]

            total_cost_heat_exchanger_FOB_after_tubes = total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]
            sensitivity_tubes = 4 * ((total_cost_heat_exchanger_FOB_after_tubes-total_cost_heat_exchanger_FOB_before_tubes)/(total_cost_heat_exchanger_FOB_before_tubes))
            sensitivity_tubes = round(sensitivity_tubes, 4)
#up to h, sensitivity is definded

            sensitivities2.append(sensitivity_tubes)
            schedule40_table = [
{"Outside diameter": 10.3,"Length":10000,"Wall thickness": 1.73,"Price": 7.31},
{"Outside diameter": 13.7,"Length":10000,"Wall thickness": 2.24,"Price": 12.66},
{"Outside diameter": 17.1,"Length":10000,"Wall thickness": 2.31,"Price": 16.85},
{"Outside diameter": 21.30,"Length":10000,"Wall thickness": 2.77,"Price": 25.32},
{"Outside diameter": 26.7,"Length":10000,"Wall thickness": 2.87,"Price":33.73},
{"Outside diameter": 33.4,"Length":10000,"Wall thickness": 3.38,"Price": 50.05},
{"Outside diameter": 42.2,"Length":10000,"Wall thickness": 3.56,"Price": 67.85},
{"Outside diameter": 48.3,"Length":10000,"Wall thickness": 3.68,"Price": 80.99},
{"Outside diameter": 60.3,"Length":10000,"Wall thickness": 3.91,"Price": 108.75},
{"Outside diameter": 73,"Length":10000,"Wall thickness": 5.16,"Price": 172.66},
{"Outside diameter": 88.9,"Length":10000,"Wall thickness": 5.49,"Price": 225.86},
{"Outside diameter": 101.6,"Length":10000,"Wall thickness": 5.74,"Price": 271.39},
{"Outside diameter": 114.3,"Length":10000,"Wall thickness": 6.02,"Price": 321.51},
{"Outside diameter": 141.3,"Length":10000,"Wall thickness": 6.55,"Price": 435.33},
{"Outside diameter": 168.3,"Length":10000,"Wall thickness": 7.11,"Price": 565.27},
{"Outside diameter": 219.1,"Length":10000,"Wall thickness": 8.18,"Price": 850.98},
{"Outside diameter": 273.0,"Length":10000,"Wall thickness": 9.27,"Price": 1205.84},
{"Outside diameter": 323.8,"Length":10000,"Wall thickness": 10.31,"Price": 1594.16},
{"Outside diameter": 355.6,"Length":10000,"Wall thickness": 11.13,"Price": 1891.02},
{"Outside diameter": 406.4,"Length":10000,"Wall thickness": 12.7,"Price": 2466.15},
{"Outside diameter": 457.0,"Length":10000,"Wall thickness": 14.27,"Price": 3116.11},
{"Outside diameter": 508.0,"Length":10000,"Wall thickness": 15.09,"Price": 3668.65},
{"Outside diameter": 610.0,"Length":10000,"Wall thickness": 17.48,"Price": 5108.51}
]


            total_cost_heat_exchanger_FOB_before_rods = total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]


            rods_table = [
            {"Outside diameter": 3.175,"Length": 10000,"Price": 1.55},
                {"Outside diameter": 6.35,"Length": 10000,"Price": 6.31},
                {"Outside diameter": 9.525,"Length":10000,"Price": 14},
                {"Outside diameter": 12.7,"Length": 10000,"Price": 24.86},
                {"Outside diameter": 15.875,"Length": 10000,"Price":38.87},
                {"Outside diameter": 19.05,"Length": 10000,"Price": 55.94},
                {"Outside diameter": 22.225,"Length": 10000,"Price": 76.17},
                {"Outside diameter": 25.4,"Length": 10000,"Price": 99.44},
                {"Outside diameter": 28.575,"Length": 10000,"Price": 125.9},
                {"Outside diameter": 31.75,"Length": 10000,"Price":155.38},
                {"Outside diameter": 34.925,"Length": 10000,"Price": 188.06},
                {"Outside diameter": 38.1,"Length": 10000,"Price": 223.74},
                {"Outside diameter": 41.275,"Length": 10000,"Price": 262.91}
                ]


            total_cost_heat_exchanger_FOB_after_rods = total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]
            sensitivity_rods = 4 * ((total_cost_heat_exchanger_FOB_after_rods-total_cost_heat_exchanger_FOB_before_rods)/(total_cost_heat_exchanger_FOB_before_rods))
            sensitivity_rods = round(sensitivity_rods, 4)
#up to h, sensitivity is definded

            sensitivities2.append(sensitivity_rods)
            print(sensitivities2)
            heat_map2.append(sensitivities2)

        print(f"Heat Map 2 is: \n\n\n {heat_map2}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Heat Map 2"
        for row_index, row_data in enumerate(heat_map2):
            for col_index, cell_value in enumerate(row_data):
                ws.cell(row=row_index+1, column=col_index+1, value=cell_value)

        wb.save("Heat map 2.xlsx")



        heat_map_matrix = np.array(heat_map)

        print(f"\n\n###################################\nheat_map_matrix is: {heat_map_matrix}")
        fig, ax = plt.subplots()
        im, cbar = heatmap(heat_map_matrix, name_examples, increased_parameters, ax=ax, cmap="YlGn", cbarlabel="Sensitivity value")
    #im, cbar = ax.imshow(heat_map_matrix)   previous
        ax.set_xticks(range(len(increased_parameters)), labels=increased_parameters,
                      rotation=-45, ha="right", rotation_mode="anchor", fontsize=9)
        ax.set_yticks(range(len(name_examples)), labels=name_examples)
        for j in range(len(increased_parameters)):
            for i in range(len(name_examples)):
                text = ax.text(j,i, heat_map_matrix[i, j],
                               ha="center", va="center", color="black")

        ax.set_title("Sensitivity of parameters for all exchangers, h=0.25")
        fig.tight_layout()
        plt.show()

        heat_map_matrix2 = np.array(heat_map2)

        fig, ax = plt.subplots()
        im, cbar = heatmap(heat_map_matrix2, name_examples, raw_material_parameters, ax=ax, cmap="YlGn", cbarlabel="Sensitivity value")
    #im, cbar = ax.imshow(heat_map_matrix)   previous
        ax.set_xticks(range(len(raw_material_parameters)), labels=raw_material_parameters,
                      rotation=-45, ha="right", rotation_mode="anchor", fontsize=9)
        ax.set_yticks(range(len(name_examples)), labels=name_examples)
        for j in range(len(raw_material_parameters)):
            for i in range(len(name_examples)):
                text = ax.text(j,i, heat_map_matrix2[i, j],
                               ha="center", va="center", color="black")

        ax.set_title("Sensitivity of raw materials for all exchangers, h=0.25")
        fig.tight_layout()
        plt.show()


      ##### make sensitivity cost for every plate first, discerning between the type of plates used
        
        """parameters2=[cost_raw_material_shell,cost_raw_material_tubesheets,cost_raw_material_tubes,cost_raw_material_heads,cost_raw_material_nozzles, cost_raw_material_flanges,cost_raw_material_removable_covers,cost_raw_material_pass_partitions,cost_raw_material_tie_rods,cost_raw_material_nozzles_flanges,cost_raw_material_baffles]
        heat_map2=[]

        raw_material_parameters = ["Plate of shell","Plate of tubesheets", "Pipes of tubes", "Plate of heads", "Pipe of nozzles", "Plates of flanges", "PLates of removable covers", "Plate of pass partitions", "Rods of tie rods", "Flanges of flanges nozzles", "Plates of baffles"]

        for example in examples:
            ds.value=example["ds"]
            ls.value=example["ls"]
            Nb.value=example["Nb"]
            dte.value=example["dte"]

            number_passes.value=example["number_passes"]
            Ntp.value=example["Ntp"]
            lay.value=example["lay"]
            ltp.value = example["ltp"]
            shell_side_mass_flow.value=example["shell_side_mass_flow"]
            tube_side_mass_flow.value=example["tube_side_mass_flow"]
            shell_side_fluid_density.value = example["shell_side_fluid_density"]
            tube_side_fluid_density.value = example["tube_side_fluid_density"]

            sensitivities2=[]
            def funct(plates_table, sch40_table,rods_table,nozzles_flanges_table,ts, dte, tts,dnzs, dnzt, trc,tpp,th,tfl, tb, dtr):
                dte=dte.value
                plates_and_tubes_used_price = []

                for plates in plates_table:
                    if plates["Thickness"]>=ts:
                        shell_plates_price = plates["Price"]
                        plates_and_tubes_used_price.append(shell_plates_price)
                        break
                    else:
                        continue
                for plates in plates_table:
                    if plates["Thickness"]<tts:
                        continue
                    else:
                        tubesheets_price = plates["Price"]
                        plates_and_tubes_used_price.append(tubesheets_price)
                        break

                for tubes in sch40_table:
                    if tubes["Outside diameter"]>=dte:
                        tubes_price = tubes["Price"]
                        plates_and_tubes_used_price.append(tubes_price)
                        break
                    else:
                        continue

                for plates in plates_table:
                    if plates["Thickness"]>=th:
                        heads_price = plates["Price"]
                        plates_and_tubes_used_price.append(heads_price)
                        break
                    else:
                        continue
                nozzles_price=0
                for tubes in sch40_table:
                    if tubes["Outside diameter"]-2*tubes["Wall thickness"]<dnzs:
                        continue
                    else:
                        nozzles_shell_price = tubes["Price"]
                        nozzles_price+=nozzles_shell_price
                        break

                for tubes in sch40_table:
                    if tubes["Outside diameter"]-2*tubes["Wall thickness"]<dnzt:
                        continue
                    else:
                        nozzles_price+= tubes["Price"]
                        break

                    plates_and_tubes_used_price.append(nozzles_tubes_price)


                for plates in plates_table:
                    if plates["Thickness"]>=tfl:
                        flanges_price = plates["Price"]
                        plates_and_tubes_used_price.append(flanges_price)
                        break
                    else:
                        continue

                for plates in plates_table:
                    if plates["Thickness"]>=trc:
                        rem_cov_price = plates["Price"]
                        plates_and_tubes_used_price.append(rem_cov_price)
                        break
                    else:
                        continue

                for plates in plates_table:
                    if plates["Thickness"]<tpp:
                        continue
                    else:
                        pp_price = plates["Price"]
                        plates_and_tubes_used_price.append(pp_price)
                        break
                for rods in rods_table:
                    if rods["Outside diameter"]<dtr:
                        continue
                    else:
                        tie_rods_price = rods["Price"]
                        plates_and_tubes_used_price.append(tie_rods_price)
                        break
                for flanges in flanges_nozzles_table:
                    nozzles_flange_price = 0
                    if flanges["Inside diameter"]>=dnzs:
                        noz_fla_price = flanges["Price"]
                        nozzles_flange_price+=noz_fla_price
                        
                    else:
                        if flanges["Price"] == 428.48:
                            nozzles_flange_price += 428.48
                        else:
                            continue
                    if flanges["Inside diameter"]>=dnzt:
                        noz_fla_price = flanges["Price"]
                        nozzles_flange_price+=noz_fla_price
                    else:
                        if flanges["Price"] == 428.48:
                            nozzles_flange_price += 428.48
                        else:
                            continue
                    plates_and_tubes_used_price.append(nozzles_flange_price)
                for plates in plates_table:
                    if plates["Thickness"]<=tb:
                        continue
                    else:
                        baf_price = plates["Price"]
                        plates_and_tubes_used_price.append(baf_price)
                        break

                return plates_and_tubes_used_price"""

        """plates_and_tubes_used_price = funct(plates_table, schedule40_table, rods_table,flanges_nozzles_table, ts, dte, tubesheets_thickness, shell_nozzles_inside_diameter,tubes_nozzles_inside_diameter, removable_covers_thickness, pass_partitions_thickness,heads_thickness,flanges_thickness,baffles_thickness, diameter_tie_rods)
            #print(f"plates and tubes  used: {plates_and_tubes_used_price}")


            #def funct(plates_table, sch40_table,rods_table,nozzles_flanges_table,ts, dte, tts,dnzs, dnzt, trc,tpp,th,tfl, tb,dtr):
            for i,parameters in enumerate(parameters2):
                original_value_total_cost_ex = total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]                        
                factor_povery = parameters/original_value_total_cost_ex
                print(f"Factor p over y is: {factor_povery}")
                original_value_parameter=parameters
                global sensitivity_check
                sensitivity_check=True
                total_cost_ex_aumented =total_cost_heat_exchanger_FOB(ds, ls, dte, lay, ltp, Nb, BC, number_passes, Ntp, max_allowable_stress, Sb, gasket_material_factor,welding_efficiency, corrosion_allowance, shell_side_pressure, tube_side_pressure, shell_side_fluid_velocity, tube_side_fluid_velocity, shell_side_mass_flow, tube_side_mass_flow, shell_side_fluid_density, tube_side_fluid_density, labor_cost_assembly, starting_time_assembly, manipulation_time_assembly, labor_cost_bolted_joints, material_cost_bolted_joints,utility_cost_bolted_joints, depreciation_cost_bolted_joints, STba, k_scrap_factor_shell, k_scrap_factor_tubes, k_scrap_factor_tubesheets, k_scrap_factor_heads, k_scrap_factor_shell_nozzles, k_scrap_factor_tubes_nozzles, k_scrap_factor_baffles, k_scrap_factor_flanges, k_scrap_factor_removable_covers, k_scrap_factor_tie_rods, k_scrap_factor_pass_partitions, k_scrap_factor_nozzles_flanges,labor_cost_plasma_cutting, material_cost_plasma_cutting, utility_cost_plasma_cutting, depreciation_cost_plasma_cutting, starting_time_plasma_cutting, manipulation_time_plasma_cutting, spca, tpcs, operation_factor_plasma_cutting, ec, labor_cost_drilling, material_cost_drilling, utility_cost_drilling, depreciation_cost_drilling, starting_time_drilling, manipulation_time_drilling, Stm, rpm, fnd, STdo, SDh, labor_cost_lathe, material_cost_lathe, utility_cost_lathe, depreciation_cost_lathe, starting_time_lathe, manipulation_time_lathe, operation_factor_lathe, vcl, apl, fnl, labor_cost_root_pass_welding, material_cost_root_pass_welding, utility_cost_root_pass_welding, depreciation_cost_root_pass_welding, starting_time_root_pass_welding, manipulation_time_root_pass_welding, density_filler_material, nrp, operation_factor_root_pass_welding, rrp, thickness_root_pass_welding, labor_cost_welding, material_cost_welding, utility_cost_welding, depreciation_cost_welding, starting_time_welding, manipulation_time_welding, operation_factor_welding, nw, rw, labor_cost_blade_saw_cutting, material_cost_blade_saw_cutting, utility_cost_blade_saw_cutting, depreciation_cost_blade_saw_cutting, starting_time_blade_saw_cutting, manipulation_time_blade_saw_cutting, Crbs, operation_factor_blade_saw_cutting, plates_table, schedule40_table, rods_table)[0]                        
                sensit_new = factor_povery *((total_cost_ex_aumented-original_value_of_total_cost)/(parameters*0.25))
                sensitivities2.append(sensit_new)
                sensitivity_check = False 
            heat_map2.append(sensitivities2)


        heat_map_matrix2 = np.array(heat_map2)
        print(f"\n\n###################################\nheat_map_matrix is: {heat_map_matrix2}")
        fig, ax = plt.subplots()
        im, cbar = heatmap(heat_map_matrix2, name_examples,raw_material_parameters, ax=ax, cmap="YlGn", cbarlabel="Sensitivity value")
        ax.set_xticks(range(len(raw_material_parameters)), labels=raw_material_parameters,
                      rotation=-45, ha="right", rotation_mode="anchor")
        ax.set_yticks(range(len(name_examples)), labels=name_examples)
        #for j in range(len(increased_parameters)):
        #    for i in range(len(name_examples)):
        #        text = ax.text(j,i, heat_map_matrix[i, j],
        #                       ha="center", va="center", color="w")

        ax.set_title("Sensitivity of raw material for all exchangers, h=0.25")
        fig.tight_layout()
        plt.show()"""











    data_entry_Row = ft.Row(
        controls=[
            ft.Column(controls=[
                data_entry,
                ds,
                ls,
                dte,
                lay,
                ltp,
                Nb,
                BC,
                number_passes,
                Ntp,
                ft.ElevatedButton("Calculate total cost",on_click=calculate_heat_exchanger_FOB_button),
                ft.ElevatedButton("Calculate sensitivity of parameter", on_click=heat_map_make)
                ]
                      ),
            ft.Column(controls=[
                ft.Container(
                    ft.Text("Parameters"),
                    bgcolor=ft.colors.RED_400,
                    padding=5,
                    border_radius=10,
                    width=200,
                    height=100,
                    alignment=ft.alignment.center
                    ),
                                    
                    ft.Container(ft.Text("Fluid parameters"), bgcolor=ft.colors.RED, padding=5, border_radius=10, width=200, height=100, alignment=ft.alignment.center),
                    ft.Container(ft.Text("Assembly settings"), bgcolor=ft.colors.RED, padding=5, border_radius=10, width=200, height=100, alignment=ft.alignment.center),
                    ft.Container(ft.Text("Economic model parameters"), bgcolor=ft.colors.RED, padding=5, border_radius=10, width=200, height=100, alignment=ft.alignment.center),
                    ft.Container(ft.Text("Operational settings"), bgcolor=ft.colors.RED, padding=5, border_radius=10, width=200, height=100, alignment=ft.alignment.center)
                ]
                      ),
            ft.Column(
                controls=[
                    ft.Container(
                        ft.Row(controls=[
                            ft.Text("Parameters"),
                            max_allowable_stress,
                            Sb,
                            gasket_material_factor,
                            welding_efficiency
                            ],
                               wrap=True
                           ),
                        border=ft.border.all(2, ft.colors.BLUE_200),
                                 ),

                    ft.Container(
                        ft.Row(controls=[
                            ft.Text("Assembly Settings"),
                            
                            labor_cost_assembly,
                            manipulation_time_assembly, 
    
                            labor_cost_bolted_joints,
                            material_cost_bolted_joints,
                            utility_cost_bolted_joints,
                            depreciation_cost_bolted_joints,
                            STba
                            ],
                               wrap=True
                               )
                        ),
                    
                    ft.Container(
                        ft.Row(
                            controls=[
                                ft.Text("Fluid Parameters"),
                                shell_side_pressure,
                                tube_side_pressure,
                                shell_side_fluid_velocity,
                                tube_side_fluid_velocity,
                                shell_side_mass_flow,
                                tube_side_mass_flow,
                                shell_side_fluid_density,
                                tube_side_fluid_density
                                ],
                                   wrap=True
                                   ),
                         border=ft.border.all(2, ft.colors.BLUE_200)
                        ),
                    
                    ft.Container(
                        ft.Row(
                            controls=[
                                ft.Text("Economic Model Parameters"),
                                cost_shipping,
                                cost_foundation,
                                km_shipping, 
                                kg_price,
                                labor_cost_LAP,
                                manipulation_time_assembly,
                                starting_time_LAP,
                                k_scrap_factor_shell,
                                k_scrap_factor_tubesheets,
                                k_scrap_factor_tubes,
                                k_scrap_factor_heads,
                                k_scrap_factor_shell_nozzles,
                                k_scrap_factor_tubes_nozzles,
                                k_scrap_factor_baffles,
                                k_scrap_factor_flanges,
                                k_scrap_factor_removable_covers,
                                k_scrap_factor_tie_rods,
                                k_scrap_factor_pass_partitions,
                                k_scrap_factor_nozzles_flanges
                                ],
                               wrap=True
                            ),
                        border=ft.border.all(2, ft.colors.BLUE_200),
                        alignment=ft.alignment.center
                        ),

                    ft.Container(
                        ft.Row(controls=[
                            ft.Text("Operational Settings"),
                            ft.Text("Plasma Cutting"),
    #Plasma Cutting operation parameters
                            labor_cost_plasma_cutting, 
                            material_cost_plasma_cutting, 
                            utility_cost_plasma_cutting, 
                            depreciation_cost_plasma_cutting, 

    #Drilling operation parameters
                            labor_cost_drilling, 
                            material_cost_drilling, 
                            utility_cost_drilling, 
                            depreciation_cost_drilling, 
    
    #Lathe operation parameters
                            labor_cost_lathe, 
                            material_cost_lathe, 
                            utility_cost_lathe, 
                            depreciation_cost_lathe, 

    #Root pass welding operation parameters
                            labor_cost_root_pass_welding, 
                            material_cost_root_pass_welding, 
                            utility_cost_root_pass_welding, 
                            depreciation_cost_root_pass_welding, 

    #Welding operation parameters
                            labor_cost_welding, 
                            material_cost_welding, 
                            utility_cost_welding, 
                            depreciation_cost_welding, 
    
    #Blade Saw Cutting operation parameters
                            labor_cost_blade_saw_cutting, 
                            material_cost_blade_saw_cutting, 
                            utility_cost_blade_saw_cutting, 
                            depreciation_cost_blade_saw_cutting
                            ],
                                  wrap=True
                                  )
                        )
                    ]
                )
            ],
            wrap=True
        )
    
    #Calculated Geometry

    data_entry_list = [ds.value, ls.value, dte.value, lay.value, ltp.value, Nb.value, BC.value, number_passes.value, Ntp.value]
    

    page.add(data_entry_Row)
    
ft.app(main)

